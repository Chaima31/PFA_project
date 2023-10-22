import os
import tabula
import pandas as pd
import datetime
from django.shortcuts import render
from django.http import HttpResponse
from openpyxl import Workbook
from .models import Client, bondedecommandee

#Acceuil page !!
from django.shortcuts import render
def ma_vue(request):
    return render(request, 'Myapp/acceuil.html')

def convert_pdf_to_excel_express(request):
    error_messages = []
    error_message = []
    combined_df = pd.DataFrame()  # Initialize an empty DataFrame to combine data
    error_df = pd.DataFrame(columns=["Error Type", "Value"])
    # Define the desired order of columns
    desired_order = ["Type doc vte", "Organis_commerciale", "Canal_distribution", "Secteur_activite",
                     "Agence_commerciale", "Groupe_Vendeur", "Client", "Receptionneur", "No commande", "Date",
                     "Articlee",
                     "Nnméro Article du client", "Qté", "Itinéraire", "DIVISION", "Poids", "TXLINE", "Néquipement"]

    if request.method == "POST" and request.FILES.getlist("pdf_file"):
        # Get the list of uploaded PDF files from the request
        pdf_files = request.FILES.getlist("pdf_file")

        for pdf_file in pdf_files:
            numero, nom = read_pdf_and_get_no_commande_express(pdf_file)
            print(f"No commande: {numero}")
            print(f"Nom de client: {nom}")

            matching_clients = Client.objects.filter(Nom_Client=nom)

            if matching_clients.exists():
                client = matching_clients.first()
                print("Matching client details:")
                print("Client:", client.Code_Client)
                print("Groupe_Vendeur:", client.Groupe_Vendeur)
                print("Designation_Vendeur:", client.Designation_Vendeur)
                print("Organis_commerciale:", client.Organis_commerciale)
                print("Canal_distribution:", client.Canal_distribution)
                print("Secteur_activite:", client.Secteur_activite)
                print("Conditions_paiement:", client.Conditions_paiement)
                print("Agence_commerciale:", client.Agence_commerciale)
                print("DIVISION:", client.DIVISON)
                print("---")
            else:
                print("No matching client found.", nom)
                error_message.append(f"No matching client found: {nom}")
                raise ValueError(error_message)

            try:
                dfs = tabula.read_pdf(pdf_file, pages="all", multiple_tables=True)
                filtered_dfs = [df for df in dfs if df.columns[0].startswith("N° article")]

                if not filtered_dfs:
                    error_messages.append("No tables with 'N° article' found in the PDF.")
                    continue

                desired_columns = ["EAN principal", "Qté"]
                filtered_dfs = [df[desired_columns] for df in filtered_dfs]
                df_merged = pd.concat(filtered_dfs)
                df_merged.dropna(subset=["Qté"], inplace=True)

                new_column_values = []
                no_commande_values = [numero] * len(df_merged)

                for ean_value in df_merged["EAN principal"]:
                    ean_str = int(ean_value)
                    matched_entry = bondedecommandee.objects.filter(Code_Barre=ean_str).first()

                    if matched_entry:
                        print("Correspondance trouvée pour EAN principal:", ean_value)
                        print("ID de l'entrée correspondante:", matched_entry.id)
                        print("Code_Barre correspondant:", matched_entry.Code_Barre)
                        print("Code_Article_magasin correspondant:", matched_entry.Code_Article_magasin)
                        print("UV:", matched_entry.UV)
                        print("---")

                        new_column_values.append(matched_entry.Code_Article_magasin)
                    else:
                        new_column_values.append(ean_value)
                        error_message.append(f"Code_Barre not found in database: {ean_value}")
                        raise ValueError(error_message)

                today_date = datetime.date.today().strftime("%Y.%m.%d")
                df_merged["Articlee"] = new_column_values
                df_merged["No commande"] = no_commande_values
                df_merged["Date"] = today_date
                df_merged["Nnméro Article du client"] = ""
                df_merged["Itinéraire"] = ""
                df_merged["Poids"] = ""
                df_merged["TXLINE"] = ""
                df_merged["Néquipement"] = ""

                code_client_values = [client.Code_Client] * len(df_merged)
                groupe_vendeur_values = [client.Groupe_Vendeur] * len(df_merged)
                canal = [client.Canal_distribution] * len(df_merged)
                secteur = [client.Secteur_activite] * len(df_merged)
                agence = [client.Agence_commerciale] * len(df_merged)
                organi = [client.Organis_commerciale] * len(df_merged)
                division = [client.DIVISON] * len(df_merged)

                df_merged["Client"] = code_client_values
                df_merged["Receptionneur"] = code_client_values
                df_merged["Groupe_Vendeur"] = groupe_vendeur_values
                df_merged["Canal_distribution"] = canal
                df_merged["Secteur_activite"] = secteur
                df_merged["Agence_commerciale"] = agence
                df_merged["Organis_commerciale"] = organi
                df_merged["DIVISION"] = division
                df_merged["Typeposte"] = ""
                df_merged["Type doc vte"] = "ZNCS"

                # Concatenate the current DataFrame with the combined DataFrame
                combined_df = pd.concat([combined_df, df_merged], ignore_index=True)

            except Exception as e:
                error_message
        if error_message and not combined_df.empty:
            # Check if there are errors and combined_df is not empty

            try:
                # Generate a unique Excel filename
                excel_filename = f"combined_data_{datetime.datetime.now().strftime('%Y%m%d%H%M%S')}.xlsx"

                # Reorder the columns based on the desired order
                combined_df = combined_df[desired_order]

                # Prepare the response to download the combined Excel file
                with pd.ExcelWriter(excel_filename, engine="openpyxl") as writer:
                    combined_df.to_excel(writer, index=False)

                # Generate a unique Notepad filename
                notepad_filename = f"notepad_data_{datetime.datetime.now().strftime('%Y%m%d%H%M%S')}.txt"

                # Create Notepad content with the data (excluding the header)
                notepad_content = combined_df.to_csv(sep='\t', index=False, header=False)

                # Write Notepad content to the file
                with open(notepad_filename, "w") as notepad_file:
                    notepad_file.write(notepad_content)

                # Generate a unique Error Excel filename
                error_excel_filename = f"error_data_{datetime.datetime.now().strftime('%Y%m%d%H%M%S')}.xlsx"

                # Create a DataFrame containing the error messages
                error_df = pd.DataFrame({"Error Message": error_message})

                # Write the error DataFrame to the Excel file
                with pd.ExcelWriter(error_excel_filename, engine="openpyxl") as writer:
                    error_df.to_excel(writer, index=False)

                # Create a ZIP file containing the Excel, Notepad, and Error Excel files
                zip_filename = f"combined_data_{datetime.datetime.now().strftime('%Y%m%d%H%M%S')}.zip"
                with zipfile.ZipFile(zip_filename, 'w', zipfile.ZIP_DEFLATED) as zipf:
                    zipf.write(excel_filename, os.path.basename(excel_filename))
                    zipf.write(notepad_filename, os.path.basename(notepad_filename))
                    zipf.write(error_excel_filename, os.path.basename(error_excel_filename))

                # Prepare the response to download the ZIP file
                with open(zip_filename, "rb") as zip_file:
                    zip_content = zip_file.read()

                zip_response = HttpResponse(zip_content, content_type="application/zip")
                zip_response["Content-Disposition"] = f"attachment; filename={zip_filename}"

                # Clean up temporary files
                os.remove(excel_filename)
                os.remove(notepad_filename)
                os.remove(error_excel_filename)

                return zip_response
            # Zip pour erreurs succes

            except Exception as e:
                error_messages.append(f"Error generating ZIP file: {e}")
        if error_message:
            # Créez un DataFrame contenant les messages d'erreur
            error_df = pd.DataFrame({"Error Message": error_message})

            # Générez un nom de fichier Excel pour les erreurs
            error_excel_filename = f"error_data_{datetime.datetime.now().strftime('%Y%m%d%H%M%S')}.xlsx"

            # Écrivez le DataFrame d'erreurs dans le fichier Excel
            with pd.ExcelWriter(error_excel_filename, engine="openpyxl") as writer:
                error_df.to_excel(writer, index=False)

            # Préparez la réponse pour télécharger le fichier Excel d'erreurs
            with open(error_excel_filename, "rb") as error_excel_file:
                error_excel_content = error_excel_file.read()

            error_excel_response = HttpResponse(error_excel_content,
                                                content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
            error_excel_response["Content-Disposition"] = f"attachment; filename={error_excel_filename}"

            # Supprimez le fichier Excel temporaire
            os.remove(error_excel_filename)

            return error_excel_response
            # zip erreurs


        elif not combined_df.empty:
            try:
                # Generate a unique Excel filename
                excel_filename = f"combined_data_{datetime.datetime.now().strftime('%Y%m%d%H%M%S')}.xlsx"

                # Reorder the columns based on the desired order
                combined_df = combined_df[desired_order]

                # Prepare the response to download the combined Excel file
                with pd.ExcelWriter(excel_filename, engine="openpyxl") as writer:
                    combined_df.to_excel(writer, index=False)

                # Generate a unique Notepad filename
                notepad_filename = f"notepad_data_{datetime.datetime.now().strftime('%Y%m%d%H%M%S')}.txt"

                # Create Notepad content with the data (excluding the header)
                notepad_content = combined_df.to_csv(sep='\t', index=False, header=False)

                # Write Notepad content to the file
                with open(notepad_filename, "w") as notepad_file:
                    notepad_file.write(notepad_content)

                # Create a ZIP file containing the Excel and Notepad files
                zip_filename = f"combined_data_{datetime.datetime.now().strftime('%Y%m%d%H%M%S')}.zip"
                with zipfile.ZipFile(zip_filename, 'w', zipfile.ZIP_DEFLATED) as zipf:
                    zipf.write(excel_filename, os.path.basename(excel_filename))
                    zipf.write(notepad_filename, os.path.basename(notepad_filename))

                # Prepare the response to download the ZIP file
                with open(zip_filename, "rb") as zip_file:
                    zip_content = zip_file.read()

                zip_response = HttpResponse(zip_content, content_type="application/zip")
                zip_response["Content-Disposition"] = f"attachment; filename={zip_filename}"

                # Clean up temporary files
                os.remove(excel_filename)
                os.remove(notepad_filename)

                return zip_response
            # zip succes

            except Exception as e:
                error_messages.append(f"Error generating ZIP file: {e}")

    else:
        return render(request, "Myapp/upload_pdf.html")
    return render(request, "Myapp/upload_pdf.html")


def read_pdf_and_get_no_commande_maxilv(pdf_file):
    # Read PDF File and Extract Text using OCR (Tesseract)
    with pdfplumber.open(pdf_file) as pdf:
        first_page = pdf.pages[0]
        text = first_page.extract_text()

        # Split the text into lines
        lines = text.splitlines()
        # Print the extracted text for inspection
        print("Extracted Text:")
        print(text)
        # Check if there are at least 5 lines in the text
        if len(lines) >= 5:
            # Get the content of the first column from the fifth line as "No de commande"
            no_commande = lines[6].split()[0]

            # Get the content of the first column from the fifth line as "Nom de client"
            nom_client = lines[8].split()[0]

            return no_commande, nom_client

        else:
            print("No commande or Nom de client not found in the PDF.")
            return None, None


def indexx(request):
    if request.method == "POST":
        pdf_files = request.FILES.getlist("pdf_file")

        if pdf_files:
            media_dir = os.path.join(settings.MEDIA_ROOT, "uploads")
            if not os.path.exists(media_dir):
                os.makedirs(media_dir)

            excel_files = []
            errors = {}
            success_files = []
            for pdf_file in pdf_files:
                pdf_file_path = os.path.join(media_dir, pdf_file.name)

                with open(pdf_file_path, 'wb') as destination:
                    for chunk in pdf_file.chunks():
                        destination.write(chunk)

                try:
                    excel_file_path = os.path.join(media_dir, f"{os.path.splitext(pdf_file.name)[0]}.xlsx")
                    excel_file_path = convert_pdf_to_excel_MAXILV(pdf_file_path, excel_file_path)
                    excel_files.append(excel_file_path)
                    success_files.append(pdf_file.name)
                except ValueError as e:
                    error_message = str(e)
                    errors[pdf_file.name] = error_message

            # Si tous les fichiers PDF ont échoué
            if not success_files and errors:
                # Créer un fichier Excel pour stocker les erreurs
                error_excel = openpyxl.Workbook()
                error_sheet = error_excel.active
                # Add headers to the new sheet
                error_sheet.append(['pdf_name', 'Remarque Client/Article'])
                # Écrire les erreurs dans le fichier Excel
                # Écrire les erreurs dans le fichier Excel
                for i, (file_name, error_message) in enumerate(errors.items(), start=1):
                    error_sheet.cell(row=i + 1, column=1, value=file_name)
                    error_sheet.cell(row=i + 1, column=2, value=error_message)

                # Sauvegarder le fichier Excel
                error_report_path = os.path.join(media_dir, 'error_report.xlsx')
                error_excel.save(error_report_path)

                # Vous pouvez maintenant retourner ce fichier Excel comme réponse
                with open(error_report_path, 'rb') as excel_file:
                    response = HttpResponse(excel_file.read(),
                                            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                    response['Content-Disposition'] = 'attachment; filename=error_report.xlsx'
                    return response

            # Combine Excel files into a single Excel file
            combined_excel_file_path = os.path.join(media_dir, "combined_excel.xlsx")
            combine_excel_files(excel_files, combined_excel_file_path)

            if errors and success_files:
                # Ouvrir le fichier Excel combiné
                combined_excel = openpyxl.load_workbook(combined_excel_file_path)

                # Supprimer la feuille vide si elle existe
                if 'Sheet' in combined_excel.sheetnames:
                    sheet_to_remove = combined_excel['Sheet']
                    combined_excel.remove(sheet_to_remove)

                # Créer une nouvelle feuille dans combined_excel.xlsx pour copier les données de error_report.xlsx
                new_sheet = combined_excel.create_sheet(title='error_report_data')

                # Ajouter les en-têtes à la nouvelle feuille
                new_sheet.append(['pdf_name', 'Remarque Client/Article'])

                # Ouvrir le fichier error_report.xlsx
                error_report_path = os.path.join(media_dir, 'error_report.xlsx')
                error_report = openpyxl.load_workbook(error_report_path)
                error_report_sheet = error_report.active

                # Copier les données de error_report.xlsx vers la nouvelle feuille
                for row in error_report_sheet.iter_rows(min_row=1, values_only=True):
                    new_sheet.append(row)

                # Ajouter les erreurs des fichiers PDF qui ont échoué à la nouvelle feuille
                for pdf_file, error_message in errors.items():
                    new_sheet.append([pdf_file, error_message])

                # Sauvegarder le fichier Excel combiné (sans la feuille vide)
                combined_excel.save(combined_excel_file_path)

                # Générer le contenu du fichier texte à partir du fichier Excel combiné
                text_content = generate_text_content(combined_excel_file_path)

                # Créer un fichier ZIP
                zip_buffer = io.BytesIO()
                with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zipf:
                    # Ajouter le fichier Excel combiné au ZIP
                    zipf.write(combined_excel_file_path, arcname='combined_excel_with_errors.xlsx')

                    # Ajouter le fichier error_report.xlsx au ZIP
                    zipf.write(error_report_path, arcname='error_report.xlsx')

                    # Ajouter le fichier texte au ZIP
                    zipf.writestr('combined_blocnote.txt', text_content)

                # Préparer la réponse HTTP pour le fichier ZIP
                response = HttpResponse(zip_buffer.getvalue(), content_type='application/zip')
                response['Content-Disposition'] = 'attachment; filename=combined_files.zip'

                return response

            elif success_files:

                if remove_empty_sheet(combined_excel_file_path):
                    # Générer le contenu du fichier texte à partir du fichier Excel combiné
                    text_content = generate_text_content(combined_excel_file_path)

                    # Télécharger le fichier Excel combiné
                    excel_response = download_excel(request, combined_excel_file_path)

                    # Créer un fichier ZIP
                    zip_buffer = io.BytesIO()
                    with zipfile.ZipFile(zip_buffer, 'w') as zipf:
                        # Ajouter le fichier Excel au ZIP
                        zipf.writestr('combined_excel.xlsx', excel_response.content)
                        # Ajouter le fichier texte au ZIP
                        zipf.writestr('combined_blocnote.txt', text_content)

                    # Définir les en-têtes appropriés pour le téléchargement du fichier ZIP
                    response = HttpResponse(zip_buffer.getvalue(), content_type='application/zip')
                    response['Content-Disposition'] = 'attachment; filename=combined_files.zip'

                    return response

    return render(request, 'Myapp/indexx.html')


def convert_pdf_to_excel_MAXILV(input_pdf_file, output_excel_file):
    # Read the PDF and get the "No de commande" and "Nom de client"
    no_commande, nom_client = read_pdf_and_get_no_commande_maxilv(input_pdf_file)

    if not no_commande or not nom_client:
        print("No commande or Nom de client not found in the PDF.")
        return
    a = 0
    # Create Excel Workbook
    wb = Workbook()
    # Liste pour stocker les DataFrames des feuilles valides
    valid_dfs = []
    # Initialize an empty list to store errors
    errors = []
    error_message = []
    nombre_total_de_lignes = 0
    # Read PDF File and Extract Tables
    # Read PDF File and Extract Tables
    with pdfplumber.open(input_pdf_file) as pdf:

        for i, page in enumerate(pdf.pages, start=1):
            table = page.extract_table()
            if table:
                header = table[0]  # Get the header row
                if "Article" in header:
                    combined_data = table[1:]  # Exclude header row
                    df = pd.DataFrame(combined_data, columns=header)
                    # Convert "Article" column to string type
                    df["Article"] = df["Article"].astype(str)

                    # Create a new column "UV" to store the corresponding UV value
                    df["UV"] = None

                    # Match "Article" column with "Code_Barre" in the bondedecommande table
                    articles_without_match = []
                    error_article = []
                    for index, row in df.iterrows():
                        article = row["Article"]
                        code_barre_id = bondedecommandeee.objects.filter(Code_Barre=article).values_list('id',
                                                                                                         flat=True).first()

                        if code_barre_id is not None:
                            code_article_magasin_id = bondedecommandeee.objects.filter(id=code_barre_id).values_list(
                                'Code_Article_magasin', flat=True).first()
                            uv = bondedecommandeee.objects.filter(Code_Barre=article).values_list('UV',
                                                                                                  flat=True).first()

                            df.at[index, "Article"] = code_article_magasin_id

                            print(
                                f"Article: {article} | Code_Barre ID: {code_barre_id} | Code_Article_magasin ID: {code_article_magasin_id}")
                        else:
                            print(article)

                            error_message.append("Code_Barre not found in database")
                            error_message.append(article)
                            raise ValueError(error_message)

                    print("Articles without a match in the database:")
                    print(articles_without_match)

                    valid_dfs.append(df)

                else:
                    print(f"Table {i} does not contain the required columns.")

        # Après avoir traité tous les PDF, imprimez le nombre total de lignes
        # Ajouter le DataFrame à la liste des DataFrames valides

        #        print(df, "sanaae")
        # Save the DataFrame to the Excel sheet
        sheet_name = f"Table_{i}"
        ws = wb.create_sheet(title=sheet_name)
        for r_idx, row in enumerate(dataframe_to_rows(df, index=False), start=1):
            for c_idx, value in enumerate(row, start=1):
                ws.cell(row=r_idx, column=c_idx, value=value)

    # ...

    # Save the Excel Workbook
    wb.save(output_excel_file)
    dfs_to_concat = []  # List to store valid DataFrames for concatenation

    # Read the Excel file again to get the valid sheet names
    valid_sheet_names = [f"Table_{j}" for j in range(1, len(valid_dfs) + 1)]

    for sheet_name in valid_sheet_names:
        try:
            df = pd.read_excel(output_excel_file, sheet_name=sheet_name)

            # Vérifier si la colonne "Article" existe
            if "Article" in df.columns:
                dfs_to_concat.append(df)
            else:
                print(f"La feuille {sheet_name} ne contient pas la colonne 'Article' requise.")
        except Exception as e:
            print(f"Erreur lors de la lecture de la feuille {sheet_name} : {e}")

        # ...

        # Rest of your code for concatenating DataFrames and saving the final DataFrame

        if valid_dfs:
            combined_df = pd.concat(valid_dfs, ignore_index=True)
            combined_df = combined_df[~combined_df["Article"].astype(str).str.contains("Article")]
            print("sanaa", combined_df)
            # Assuming you have the combined_df DataFrame somewhere in your code
            # generate_report(combined_df)

            # Enregistrer le DataFrame combiné dans le fichier Excel
            combined_df.to_excel(output_excel_file, sheet_name="Combined_Table", index=False)
            print(combined_df.shape[0], "nombre de ligne de table article de ce pdf selectionner")
        else:
            print("Aucune feuille valide trouvée pour la concaténation.")

    print("All tables combined and saved to Excel.")
    print(f"No commande: {no_commande}")
    print(f"nom de client: {nom_client}")

    # Variable to store the id of the corresponding client Quanten\nUVC
    matching_client_id = None

    clients = Client.objects.all()

    # Loop through the records and print each column's value
    for client in clients:
        # Remove spaces from the "Nom_Client" field of the current client object
        nom_client_stripped = client.Nom_Client.replace(" ", "")

        # Compare the stripped Nom_Client from the PDF with the stripped Nom_Client from the database
        if nom_client_stripped.lower() == nom_client.lower():
            print("Comparison Result: True (Match)")
            # Save the id of the corresponding client
            matching_client_id = client.id

            # Display the information of the corresponding client
            print(f"ID du client correspondant: {matching_client_id}")
            print(f"Groupe_Vendeur: {client.Groupe_Vendeur}")
            print(f"Designation_Vendeur: {client.Designation_Vendeur}")
            print(f"Conditions_paiement: {client.Conditions_paiement}")
            print(f"Agence_commerciale: {client.Agence_commerciale}")
            print(f"Nom: {client.Nom_Client}")
            print(f"DIVISON: {client.DIVISON}")
            print("\n")

            # Exit the loop as we have found a match
            break

    # Check if a matching client has been found
    if matching_client_id:
        print(f"Le client correspondant a été trouvé.")

        # Add the "Groupe_Vendeur", "Designation_Vendeur", and "Conditions_paiement" columns to the "Combined_Table" DataFrame
        import datetime

        combined_df.insert(0, "Type doc vte", "ZNCS")
        combined_df.insert(1, "Organis.commerciale", "V500")
        combined_df.insert(2, "Canal_distribution", "BG")
        combined_df.insert(3, "Secteur_activite", "ZN")
        combined_df.insert(4, "Agence", client.Agence_commerciale)
        combined_df.insert(5, "Groupe de vendeurs", client.Groupe_Vendeur)

        import datetime
        today_date = datetime.date.today().strftime("%Y.%m.%d")
        # combined_df.insert(5, "Nom", client.Nom_Client)
        current_date = datetime.date.today().strftime("%Y-%m-%d")
        combined_df.insert(6, "Donneur", client.Code_Client)
        combined_df.insert(7, " Rece", client.Code_Client)
        combined_df.insert(8, "N", no_commande)
        combined_df.insert(9, "Date", today_date)
        combined_df.insert(11, "Num", "")
        combined_df.insert(13, "Itinéraire", "")
        combined_df.insert(14, "Division", client.DIVISON)
        combined_df.insert(15, "Poids", "")
        combined_df.insert(16, "TXLINE_02", "")
        combined_df.insert(17, "N équipement", "")
        # combined_df.insert(9, " Article", article)
        # combined_df.insert(10, "unité de vente", uv)
        # combined_df.insert(11, "Type poste", "TAN")
        # combined_df.insert(12, "DIVISION", client.DIVISON)

        # Save the updated DataFrame to the Excel file with the new columns
        combined_df.to_excel(output_excel_file, sheet_name="Combined_Table", index=False)

        # ...
        # combined_df["Type poste"] ="TAN"
        # combined_df["DIVISON"] = client.DIVISON

        # combined_df["Désignation"] = client.Designation_Vendeur
        # combined_df["Conditions paiement"] = client.Conditions_paiement

        # Remove the columns VL and Noligne from the DataFrame
        combined_df.drop(
            columns=["VL", "Noligne", "TypeU.C.", "UVC/UC", "UV", "Libellearticle", "Unite", "Prixachat\nnet",
                     "Totalprix\nachatnet"], inplace=True)

        # Remove columns that end with the word "speciale" from the DataFrame
        columns_to_drop = [col for col in combined_df.columns if col.lower().endswith("speciale")]
        combined_df.drop(columns=columns_to_drop, inplace=True)

        # Remove columns that end with the word "UC" from the DataFrame
        columns_to_drop_UC = [col for col in combined_df.columns if col.lower().endswith("UC")]
        combined_df.drop(columns=columns_to_drop_UC, inplace=True)
        # Remove columns that end with the word "UC" from the DataFrame
        # columns_to_drop_libelle = [col for col in combined_df.columns if col.lower().endswith("Libellearticle")]
        # combined_df.drop(columns=columns_to_drop_libelle, inplace=True)

        # Renommer la colonne "quanten" si elle existe
        column_to_replace = next((col for col in combined_df.columns if col.lower().startswith("quanten")), None)
        if column_to_replace:
            combined_df.rename(columns={column_to_replace: "quantité"}, inplace=True)

        # Déplacer la colonne "quantité d'ordre" après la colonne "Article"
        if "Num" in combined_df.columns and "quantité" in combined_df.columns:
            article_index = combined_df.columns.get_loc("Num")
            quantity_order_index = combined_df.columns.get_loc("quantité")
            if quantity_order_index != article_index + 1:
                combined_df = combined_df[combined_df.columns[:article_index + 1].tolist() +
                                          combined_df.columns[
                                          quantity_order_index:quantity_order_index + 1].tolist() +
                                          combined_df.columns[article_index + 1:quantity_order_index].tolist() +
                                          combined_df.columns[quantity_order_index + 1:].tolist()]
        first_column_to_drop_article = next((col for col in combined_df.columns if col.lower().startswith("Num")),
                                            None)
        if first_column_to_drop_article:
            combined_df.drop(columns=first_column_to_drop_article, inplace=True)

            # column_to_replace_libelle = next((col for col in combined_df.columns if col.lower().startswith("Libelle article")), None)
            # if column_to_replace_libelle:
            #     combined_df.rename(columns={column_to_replace_libelle: "Désignation article"}, inplace=True)
        # Save the updated DataFrame to the Excel file with the new columns
        # Renommer la colonne "Article" en "Articlee"
        if "Article" in combined_df.columns:
            combined_df.rename(columns={"Article": "Articlee"}, inplace=True)
        combined_df.to_excel(output_excel_file, sheet_name="Combined_Table", index=False)
    else:
        print(f"Erreur:'{nom_client}'")
        # error_message_client = f"Le client n'a pas été trouvé dans la base de données pour '{nom_client}'"
        error_message.append("Nom client not found")
        error_message.append(nom_client)
        raise ValueError(error_message)

        # Call the generate_error_excel function to create an error Excel file

    return output_excel_file


def read_pdf_and_get_no_commande_express(pdf_file):
    # Read PDF File and Extract Text using OCR (Tesseract)
    with pdfplumber.open(pdf_file) as pdf:
        first_page = pdf.pages[0]
        text = first_page.extract_text()

        # Split the text into lines
        lines = text.splitlines()
        # Print the extracted text for inspection
        print("Extracted Text:")
        print(text)
        # Check if there are at least 5 lines in the text
        if len(lines) >= 16:
            # Get the content of the first column from the fifth line as "No de commande"
            no_commande = lines[12].split()[2]

            # Get the content of the first, second, and third columns from the third line as "Nom de client"
            nom_client = ' '.join(lines[3].split()[0:3])

            return no_commande, nom_client

        else:
            print("No commande or Nom de client not found in the PDF.")
            return None, None


from django.contrib.auth.decorators import login_required
from django.shortcuts import render

# Reste du code reste inchangé
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font

import tabula
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font


def convert_pdf_to_excel_generic(input_pdf_file, output_excel_file):
    # Read the PDF and get the "No de commande" and "Nom de client"

    numeroC, ClientCarrefour = read_pdf_and_get_info(input_pdf_file)
    print(f"No commande: {numeroC}")
    print(f"Nom de client: {ClientCarrefour}")

    # Variable to store the id of the corresponding client
    matching_client_id = None

    clients = Client.objects.all()
    # Loop through the records and print each column's value
    # print(f"Extracted 'Nom de client' from PDF: {ClientCarrefour}")

    for client in clients:
        clientcarrefor = []
        nom_client_stripped = client.Nom_Client.replace(" ", "")
        # print(f"Comparing with 'Nom_Client' in the database: {nom_client_stripped}")
        # Compare the stripped Nom_Client from the PDF with the stripped Nom_Client from the database
        if nom_client_stripped.lower() == ClientCarrefour.lower():
            print("Comparison Result: True (Match)")
            # Save the id of the corresponding client
            matching_client_id = client.id

            # Display the information of the corresponding client
            print(f"ID du client correspondant: {matching_client_id}")
            print(f"Groupe_Vendeur: {client.Groupe_Vendeur}")
            print(f"Designation_Vendeur: {client.Designation_Vendeur}")
            print(f"Conditions_paiement: {client.Conditions_paiement}")
            print(f"Agence_commerciale: {client.Agence_commerciale}")
            print(f"Nom: {client.Nom_Client}")
            print(f"Client: {client.Code_Client}")
            print("\n")
            # Exit the loop as we have found a match

            # Get the value of DIVISON from the matched client
            division_value = client.DIVISON
            print(f"DIVISON: {division_value}")
            # Get the values of the additional attributes from the matched client
            groupe_vendeur_value = client.Groupe_Vendeur
            designation_vendeur_value = client.Designation_Vendeur
            conditions_paiement_value = client.Conditions_paiement
            agence_commerciale_value = client.Agence_commerciale
            code_client_value = client.Code_Client
            break

    else:
        # This block will be executed if the loop completes without finding a match
        print(f"Erreur: Le client '{ClientCarrefour}' n'a pas été trouvé dans la base de données.")
        error_message = f"Erreur: Le client '{ClientCarrefour}' n'a pas été trouvé dans la base de données."
        clientcarrefor.append(ClientCarrefour)
        raise ValueError(error_message)

    print("clientcarrefour", clientcarrefor)

    # Exemple : Extraire les tables du PDF à l'aide de tabula
    tables = tabula.read_pdf(input_pdf_file, pages="all", multiple_tables=True)

    # Filtrer les tables pour ne conserver que celle qui contient le champ "Libelle article"
    # Filtrer les tables pour ne conserver que celle qui contient le champ "Libelle article"
    tables_filtrees = [table for table in tables if "Libelle article" in table.columns]

    if not tables_filtrees:
        print("Aucune table contenant le champ 'Libelle article' n'a été trouvée dans le PDF.")
        return

    # Extraire la première table qui contient le champ "Libelle article"
    table_avec_code = tables_filtrees[0]

    # Colonnes à supprimer du DataFrame
    colonnes_a_supprimer = ["Libelle article", "VL", "No ligne", "UVC/UC"]

    # Vérifier si les colonnes 'Type' et 'No.' sont présentes dans le DataFrame
    if "Type" in table_avec_code.columns and "No." in table_avec_code.columns:
        colonnes_a_supprimer.extend(["Type", "No."])  # Ajouter 'Type' et 'No.' aux colonnes à supprimer
    else:
        print("Les colonnes 'Type' et/ou 'No.' ne sont pas présentes dans le DataFrame.")

    df = table_avec_code.drop(columns=colonnes_a_supprimer)
    df["Quant en"] = table_avec_code.iloc[:, 7]  # Assuming "Quant en" is in the 8th column (index 7)
    # Enregistrer le tableau dans un DataFrame
    # df = pd.DataFrame(table_with_code)
    # Créer un classeur Excel et une feuille de calcul
    wb = Workbook()
    ws = wb.active

    # Écrire les noms de colonnes dans la première ligne de la feuille de calcul
    for c_idx, column_name in enumerate(df.columns, start=1):
        ws.cell(row=1, column=c_idx, value=column_name)
        ws.cell(row=1, column=c_idx).font = Font(bold=True)

    # Écrire les données du DataFrame à partir de la deuxième ligne de la feuille de calcul
    for r_idx, row in enumerate(df.itertuples(), start=2):
        for c_idx, value in enumerate(row[1:], start=1):
            ws.cell(row=r_idx, column=c_idx, value=value)

        # Write the values for 'Quant en' in each row
    for r_idx in range(2, len(df) + 2):  # Start at row index 2 and go until the last row
        # Write the 'Quant en' value to the new column in the Excel sheet
        ws.cell(row=r_idx, column=10, value=df.loc[r_idx - 2, "Quant en"])

    # Rename the header of the "Quant en" column to "Quantite de vente"
    ws.cell(row=1, column=10, value="quantité")

    # Sauvegarder le classeur Excel dans le fichier
    wb.save(output_excel_file)

    # Reload the workbook to read the data from the saved Excel file
    wb = load_workbook(output_excel_file)
    ws = wb.active

    # Add a new column for 'Code_Article_magasin'
    ws.insert_cols(9)  # Insert the new column as the third column (index 2)
    # Set the header for the new column
    ws.cell(row=1, column=9, value="Articlee")
    ws.cell(row=1, column=9).font = Font(bold=True)

    # ws.insert_cols(10)  # Insert the new column after the 'Code_Article_magasin' column
    # Set the header for the new column
    # ws.cell(row=1, column=10, value="UV")
    # ws.cell(row=1, column=10).font = Font(bold=True)

    # Get the data from the 'code_Barre' column and compare with the database
    code_barre_column = ws['B']  # Assuming 'code_Barre' column is in the 2nd column (index 1)
    error_messages = []
    code = []
    print("Résultats de comparaison avec la base de données:")
    for row_idx, cell in enumerate(code_barre_column, start=1):  # Start at 1 since the header is at index 0
        if cell.value is not None and cell.value != "Code EAN":  # Skip the row with "Code EAN" content
            code_barre_excel = cell.value
            matching_entry = bondedecommandeeee.objects.filter(Code_Barre=code_barre_excel).first()

            if matching_entry is not None:
                code_article_magasin = matching_entry.Code_Article_magasin
                # uv_value = matching_entry.UV
                ws.cell(row=row_idx, column=9,
                        value=code_article_magasin)  # Write 'Code_Article_magasin' to the new column
                # ws.cell(row=row_idx, column=10, value=uv_value)  # Write 'UV' to the new column  | UV: {uv_value}
                print(
                    f"Code_Barre (Excel) [{row_idx}]: {code_barre_excel} | Corresponding Code_Article_magasin: {code_article_magasin}")
            else:
                print(f"Code_Barre {code_barre_excel} | Not found in the database.")
                code.append(code_barre_excel)
                error_message = code_barre_excel
                error_messages.append(error_message)
                raise ValueError(error_message)

    # return error_messages
    from openpyxl.styles import Alignment
    # Save the modified workbook with the additional column for 'Code_Article_magasin'
    wb.save(output_excel_file)
    # Recharger le classeur pour lire les données du fichier Excel sauvegardé
    wb = load_workbook(output_excel_file)
    ws = wb.active

    # Get the count of rows with "Code EAN" from the PDF
    code_ean_count_pdf = table_avec_code["Code EAN"].count()

    # Get the count of rows in the DataFrame
    df_row_count = len(df)

    # Compare the counts and adjust the DataFrame rows if necessary
    if code_ean_count_pdf != df_row_count:
        if code_ean_count_pdf < df_row_count:
            # Remove rows from the DataFrame that are beyond the count in the PDF
            df = df.head(code_ean_count_pdf)
        else:
            # Add empty rows to the DataFrame if the PDF has more rows
            df = df.append([{}] * (code_ean_count_pdf - df_row_count), ignore_index=True)

    # Find the index of the "Code EAN" column
    code_ean_column_index = None
    for col_idx, header in enumerate(ws.iter_cols(min_row=1, max_row=1), start=1):
        if header[0].value == "Code EAN":
            code_ean_column_index = col_idx

    if code_ean_column_index is not None:
        # List to store rows that need to be deleted
        rows_to_delete = []
        stop_processing = False

        # Iterate through the rows starting from the second row (skip the header)
        for r_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            # Check if the "Code EAN" (Code_Barre) column is empty
            if not row[code_ean_column_index - 1]:
                rows_to_delete.append(r_idx)
            else:
                # Check if the "Code EAN" content is "stoppé"
                code_ean_value = str(row[code_ean_column_index - 1]).strip().lower()
                if code_ean_value == "stoppé":
                    # If "Code EAN" is "stoppé," stop processing and mark to remove the following rows
                    stop_processing = True
                    break
        else:
            # If the loop completes without finding "stoppé," remove the rows with empty "Code EAN"
            for row_idx in reversed(rows_to_delete):
                ws.delete_rows(row_idx)

        # If "Code EAN" is "stoppé," remove the rows that follow
        if stop_processing:
            rows_to_delete.extend(range(r_idx + 1, ws.max_row + 1))
            for row_idx in reversed(rows_to_delete):
                ws.delete_rows(row_idx)

        # Save the modified workbook with rows removed
        wb.save(output_excel_file)

        if rows_to_delete or stop_processing:
            print("Certaines lignes ont été supprimées car le 'Code EAN' était vide ou 'stoppé'.")
        else:
            print("Aucune ligne avec 'Code EAN' vide ou 'stoppé' n'a été trouvée.")
    else:
        print("Le nom de la colonne 'Code_Barre' n'a pas été trouvé dans le fichier Excel.")

    # Adapter la hauteur de ligne pour la colonne 6 pour gérer les contenus longs
    max_length = 0
    for cell in ws['F']:  # Colonnes contenant 'No ligne'
        try:
            if len(str(cell.value)) > max_length:
                max_length = len(cell.value)
        except:
            pass
    adjusted_width = (max_length + 2) * 1.2
    ws.column_dimensions['F'].width = adjusted_width

    # ... (reste du code)
    # ... (reste du code)

    # Add columns for 'Numero de commande' and 'Nom de client'
    ws.insert_cols(1)  # Insert the first column for 'Nom de client'
    ws.insert_cols(1)  # Insert the second column for 'Numero de commande'

    # Set the headers for the new columns
    ws.cell(row=1, column=1, value="Type doc vte")
    ws.cell(row=1, column=2, value="Organis.commerciale")
    ws.cell(row=1, column=3, value="Canal_distribution")
    ws.cell(row=1, column=4, value="Secteur_activite")
    # ws.cell(row=1, column=13, value="Type de poste")
    ws.cell(row=1, column=9, value="Ncommande")
    ws.cell(row=1, column=9).font = Font(bold=True)
    # ws.cell(row=1, column=3, value="Nom de client")
    # ws.cell(row=1, column=3).font = Font(bold=True)
    ws.cell(row=1, column=15, value="DIVISON")
    ws.cell(row=1, column=15).font = Font(bold=True)
    ws.cell(row=1, column=16, value="Poids net du poste")
    ws.cell(row=1, column=16).font = Font(bold=True)
    ws.cell(row=1, column=17, value="TXLINE_02")
    ws.cell(row=1, column=17).font = Font(bold=True)
    ws.cell(row=1, column=18, value="N équipement")
    ws.cell(row=1, column=18).font = Font(bold=True)
    ws.cell(row=1, column=5, value="Agence")
    ws.cell(row=1, column=5).font = Font(bold=True)
    # ws.cell(row=1, column=10, value="Conditions_paiement")
    # ws.cell(row=1, column=10).font = Font(bold=True)
    # ws.cell(row=1, column=11, value="Designation_Vendeur")
    # ws.cell(row=1, column=11).font = Font(bold=True)
    ws.cell(row=1, column=6, value="Groupe_Vendeur")
    ws.cell(row=1, column=6).font = Font(bold=True)
    ws.cell(row=1, column=7, value="Donneur")
    ws.cell(row=1, column=7).font = Font(bold=True)
    ws.cell(row=1, column=8, value="Réceptionneur")
    ws.cell(row=1, column=8).font = Font(bold=True)
    ws.cell(row=1, column=10, value="Date")
    ws.cell(row=1, column=10).font = Font(bold=True)
    ws.cell(row=1, column=12, value="Nnméro Article du client")
    ws.cell(row=1, column=12).font = Font(bold=True)
    ws.cell(row=1, column=14, value="Itinéraire")
    ws.cell(row=1, column=14).font = Font(bold=True)

    # Write the values for 'Numero de commande', 'Nom de client', and 'DIVISON' in each row
    for r_idx in range(2, len(df) + 2):  # Start at row index 2 and go until the last row
        ws.cell(row=r_idx, column=1, value="ZNCS")
        ws.cell(row=r_idx, column=2, value="V500")
        ws.cell(row=r_idx, column=3, value="GS")
        ws.cell(row=r_idx, column=4, value="ZN")
        # ws.cell(row=r_idx, column=13, value="TAN")
        ws.cell(row=r_idx, column=9, value=numeroC)
        # ws.cell(row=r_idx, column=3, value=ClientCarrefour)
        ws.cell(row=r_idx, column=15, value=division_value)  # Write 'DIVISON' to the new column
        ws.cell(row=r_idx, column=5, value=agence_commerciale_value)  # Write 'DIVISON' to the new column
        # ws.cell(row=r_idx, column=10, value=conditions_paiement_value)  # Write 'DIVISON' to the new column
        # ws.cell(row=r_idx, column=11, value=designation_vendeur_value)  # Write 'DIVISON' to the new column
        ws.cell(row=r_idx, column=6, value=groupe_vendeur_value)  # Write 'DIVISON' to the new column
        ws.cell(row=r_idx, column=7, value=code_client_value)  # Write 'DIVISON' to the new column
        ws.cell(row=r_idx, column=8, value=code_client_value)  # Write 'DIVISON' to the new column
        import datetime
        today_date = datetime.date.today().strftime("%Y.%m.%d")
        ws.cell(row=r_idx, column=10, value=today_date)  # Write 'DIVISON' to the new column
        ws.cell(row=r_idx, column=12, value="")  # Write 'DIVISON' to the new column
        ws.cell(row=r_idx, column=14, value="")  # Write 'DIVISON' to the new column
        ws.cell(row=r_idx, column=16, value="")  # Write 'DIVISON' to the new column
        ws.cell(row=r_idx, column=17, value="")  # Write 'DIVISON' to the new column
        ws.cell(row=r_idx, column=18, value="")  # Write 'DIVISON' to the new column

    # ...
    # Add a column for today's date
    # ws.insert_cols(1)  # Insert the first column for today's date

    # Get today's date in the desired format (e.g., "YYYY-MM-DD" or "DD-MM-YYYY")
    today_date = datetime.date.today().strftime("%Y.%m.%d")

    # Set the header for the new column
    # ws.cell(row=1, column=1, value="Date")
    # ws.cell(row=1, column=1).font = Font(bold=True)

    # Write today's date in each row of the new column
    # for r_idx in range(2, len(df) + 2):  # Start at row index 2 and go until the last row
    #    ws.cell(row=r_idx, column=1, value=today_date)
    # ...
    # ...
    # ...
    for r_idx in range(2, len(df) + 2):
        current_quantite = ws.cell(row=r_idx, column=13).value
        current_article = ws.cell(row=r_idx, column=11).value
        print(f"Ligne {r_idx}: Valeur actuelle de Quantité = {current_quantite}")
        if current_quantite is not None and isinstance(current_quantite, (int, float)) and current_article is not None:
            # Query the database to retrieve the ID based on the "Article"
            code_barre_id = bondedecommandeeee.objects.filter(Code_Article_magasin=current_article).values_list('Colisage',
                                                                                                       flat=True).first()
            new_quantite = int(current_quantite / code_barre_id)

            ws.cell(row=r_idx, column=13, value=new_quantite)
            print(f"Ligne {r_idx}: Ancienne quantité = {current_quantite}, Nouvelle quantité = {new_quantite}")

    # (Le reste du code continue ici)
    # Enregistrez le classeur Excel dans le fichier
    wb.save(output_excel_file)
    print("Enregistrement dans le fichier Excel terminé.")
    return error_messages


def read_pdf_and_get_info(pdf_file):
    # Lire le fichier PDF et extraire le texte en utilisant OCR (Tesseract)
    with pdfplumber.open(pdf_file) as pdf:
        first_page = pdf.pages[0]
        text = first_page.extract_text()

        # Diviser le texte en lignes
        lines = text.splitlines()

        # Afficher le texte extrait pour inspection
        print("Texte extrait:")
        print(text)

        # Vérifier si le texte contient au moins 10 lignes
        if len(lines) >= 10:
            # Extraire le "numeroC" à partir de la 8ème ligne, 1ère colonne
            numeroC = lines[7].split()[0]

            # Extraire "ClientCarrefour" à partir de la 10ème ligne, 1ère colonne
            ClientCarrefour = lines[9].split()[0]

            return numeroC, ClientCarrefour
        else:
            # Si le texte ne contient pas suffisamment de lignes, retourner None pour les deux valeurs
            return None, None


from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows

import os

from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import os

from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import os

from openpyxl import load_workbook
import os
from openpyxl import Workbook, load_workbook
import os
import pdfplumber  # Import the pdfplumber library


def combine_excel_filess(excel_files, combined_excel_file_path):
    combined_wb = Workbook()

    for file_index, file_path in enumerate(excel_files):
        wb = load_workbook(file_path, data_only=True)

        for sheet_name in wb.sheetnames:
            if sheet_name not in combined_wb.sheetnames:
                combined_wb.create_sheet(title=sheet_name)

            source_sheet = wb[sheet_name]
            combined_sheet = combined_wb[sheet_name]

            for row_index, row in enumerate(source_sheet.iter_rows(values_only=True), start=1):
                if file_index == 0 or row_index > 1:  # Keep header for first file, exclude for others
                    combined_sheet.append(row)

    combined_wb.save(combined_excel_file_path)


from django.http import HttpResponse, JsonResponse
import os
import openpyxl


def download_links(request):
    if request.method == "POST" and request.FILES.getlist("pdf_file"):
        pdf_files = request.FILES.getlist("pdf_file")
        excel_files_to_combine = []  # Liste pour stocker les fichiers Excel convertis
        error_messages = []  # Liste pour stocker les messages d'erreur

        for pdf_file in pdf_files:
            if pdf_file.name.lower().endswith(".pdf"):
                with open("temp_pdf.pdf", "wb") as destination:
                    for chunk in pdf_file.chunks():
                        destination.write(chunk)
                try:
                    output_excel_file = f"converted_excel_{pdf_file.name}.xlsx"
                    convert_pdf_to_excel_generic("temp_pdf.pdf", output_excel_file)
                    os.remove("temp_pdf.pdf")
                    excel_files_to_combine.append(output_excel_file)
                except ValueError as e:
                    error_message = str(e)
                    error_messages.append({'file_name': pdf_file.name, 'error_message': error_message})

        # Créer un classeur Excel pour stocker les données
        combined_workbook = openpyxl.Workbook()
        combined_workbook.remove(combined_workbook.active)  # Supprimer la première feuille créée automatiquement

        if error_messages and excel_files_to_combine:
            # Créez un fichier temporaire pour le texte à partir de combined_output.xlsx
            combined_excel_file_path = "combined_output.xlsx"
            combine_excel_files(excel_files_to_combine, combined_excel_file_path)
            text_content = generate_text_content(combined_excel_file_path)

            # Créez un fichier ZIP
            zip_buffer = io.BytesIO()
            with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zipf:
                # Ajouter le fichier Excel combiné au ZIP
                zipf.write(combined_excel_file_path, arcname='combined_output.xlsx')

                # Créer un fichier temporaire pour le fichier texte
                text_file_path = "combined_blocnote.txt"
                with open(text_file_path, 'w') as text_file:
                    text_file.write(text_content)

                # Ajouter le fichier texte au ZIP
                zipf.write(text_file_path, arcname='combined_blocnote.txt')

                # Créer un fichier Excel pour les messages d'erreur
                error_report_path = "error_report.xlsx"
                error_workbook = openpyxl.Workbook()
                error_sheet = error_workbook.active
                error_sheet.title = "error_report"

                for i, error in enumerate(error_messages, start=1):
                    error_sheet.cell(row=i, column=1, value=error['file_name'])
                    error_sheet.cell(row=i, column=2, value=error['error_message'])

                error_workbook.save(error_report_path)

                # Ajouter le fichier Excel des messages d'erreur au ZIP
                zipf.write(error_report_path, arcname='error_report.xlsx')

                # Supprimer les fichiers temporaires
                os.remove(combined_excel_file_path)
                os.remove(text_file_path)
                os.remove(error_report_path)

            # Préparer la réponse HTTP pour le fichier ZIP
            response = HttpResponse(zip_buffer.getvalue(), content_type='application/zip')
            response['Content-Disposition'] = 'attachment; filename=combined_files.zip'

            return response

        elif error_messages:
            # Si seulement des erreurs sont présentes, renvoyez le fichier Excel d'erreurs
            error_excel = openpyxl.Workbook()
            error_sheet = error_excel.active
            error_sheet.title = "error_report"

            for i, error in enumerate(error_messages, start=1):
                error_sheet.cell(row=i, column=1, value=error['file_name'])
                error_sheet.cell(row=i, column=2, value=error['error_message'])

            error_report_path = "error_report.xlsx"
            error_excel.save(error_report_path)

            response = HttpResponse(open(error_report_path, "rb").read(),
                                    content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
            response['Content-Disposition'] = f'attachment; filename="{error_report_path}"'

            os.remove(error_report_path)

            return response

        elif excel_files_to_combine:

            # Combinez les fichiers Excel en un seul fichier

            combined_excel_file_path = "combined_output.xlsx"

            combine_excel_files(excel_files_to_combine, combined_excel_file_path)

            # Générer le contenu du fichier texte à partir du fichier Excel combiné

            text_content = generate_text_content(combined_excel_file_path)

            # Créer un fichier ZIP

            zip_buffer = io.BytesIO()

            with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zipf:

                # Ajouter le fichier Excel combiné au ZIP

                zipf.write(combined_excel_file_path, arcname='combined_excel.xlsx')

                # Ajouter le fichier texte au ZIP

                zipf.writestr('combined_blocnote.txt', text_content)

            # Préparer la réponse HTTP pour le fichier ZIP

            response = HttpResponse(zip_buffer.getvalue(), content_type='application/zip')

            response['Content-Disposition'] = 'attachment; filename=combined_files.zip'

            # Supprimer les fichiers temporaires (Excel et fichier combiné)

            for excel_file in excel_files_to_combine:
                os.remove(excel_file)

            os.remove(combined_excel_file_path)

            return response

    return render(request, 'Myapp/download_links.html')


from django.shortcuts import render
from .models import Client

from django.shortcuts import render, redirect, get_object_or_404
from django.core.paginator import Paginator
from django.shortcuts import render
from .models import Client


def clients_table(request):
    clients_list = Client.objects.all()
    paginator = Paginator(clients_list, 10)  # 10 clients par page
    page = request.GET.get('page')
    clients = paginator.get_page(page)
    return render(request, 'Myapp/clients_table.html', {'clients': clients})


from .models import Report
from django.shortcuts import render


def report_table(request):
    reports_list = Report.objects.all()
    paginator = Paginator(reports_list, 10)  # 10 clients par page
    page = request.GET.get('page')
    reports = paginator.get_page(page)
    return render(request, 'Myapp/report_table.html', {'reports': reports})


# Importez le modèle Client s'il n'est pas déjà importé

from django.shortcuts import render, redirect
from django.shortcuts import render, redirect
from .models import ClientForm
from .models import bondedecommandeForm
#from .models import bondedecommandeee
from .models import bondedecommandeeee


def ajouter_client(request):
    if request.method == 'POST':
        form = ClientForm(request.POST)
        if form.is_valid():
            form.save()
            # Redirect to the clients_table URL after successful form submission
            return redirect('clients_table')
        else:
            # Print form data and errors to the console for debugging
            print(request.POST)
            print(form.errors)
            # If the form is not valid, re-render the page with the form and its errors
            return render(request, 'Myapp/ajouter_client.html', {'form': form})
    else:
        # If the request is a GET request, create an empty form to display to the user
        form = ClientForm()
        return render(request, 'Myapp/ajouter_client.html', {'form': form})


def ajouter_bondedecommande(request):
    if request.method == 'POST':
        form = bondedecommandeForm(request.POST)
        if form.is_valid():
            form.save()
            # Redirect to the clients_table URL after successful form submission
            return redirect('bondedecommande')
        else:
            # Print form data and errors to the console for debugging
            print(request.POST)
            print(form.errors)
            # If the form is not valid, re-render the page with the form and its errors
            return render(request, 'Myapp/ajouter_bondedecommande.html', {'form': form})
    else:
        # If the request is a GET request, create an empty form to display to the user
        form = bondedecommandeForm()
        return render(request, 'Myapp/ajouter_bondedecommande.html', {'form': form})


def bondedecommandes(request):
    bondedecommandes_list = bondedecommandeeee.objects.all()
    paginator = Paginator(bondedecommandes_list, 10)  # 10 clients par page
    page = request.GET.get('page')
    bondedecommandes = paginator.get_page(page)
    return render(request, 'Myapp/bondedecommande.html', {'bondedecommandes': bondedecommandes})


def edit_client(request, client_id):
    client = get_object_or_404(Client, id=client_id)

    if request.method == 'POST':
        form = ClientForm(request.POST, instance=client)
        if form.is_valid():
            form.save()
            return redirect(
                'clients_table')  # Replace 'clients_table' with the actual URL name for the client list view
    else:
        form = ClientForm(instance=client)

    return render(request, 'Myapp/edit_client.html', {'form': form})


def edit_bondedecommande(request, bondedecommande_id):
    bondedecommande = get_object_or_404(bondedecommandeeee, id=bondedecommande_id)

    if request.method == 'POST':
        form = bondedecommandeForm(request.POST, instance=bondedecommande)
        if form.is_valid():
            form.save()
            return redirect('bondedecommande')
    else:
        form = bondedecommandeForm(instance=bondedecommande)

    return render(request, 'Myapp/edit_bondedecommande.html', {'form': form})


from django.shortcuts import render, redirect, get_object_or_404
from .models import Client


def delete_client(request, client_id):
    client = get_object_or_404(Client, id=client_id)

    if request.method == 'POST':
        client.delete()
        return redirect('clients_table')  # Replace 'clients_table' with the actual URL name for the client list view

    return render(request, 'Myapp/delete_client.html', {'client': client})


def delete_bondedecommande(request, bondedecommande_id):
    bondedecommande = get_object_or_404(bondedecommandeeee, id=bondedecommande_id)

    if request.method == 'POST':
        bondedecommande.delete()
        return redirect('bondedecommande')  # Replace 'clients_table' with the actual URL name for the client list view

    return render(request, 'Myapp/delete_bondedecommande.html', {'bondedecommande': bondedecommande})


import openpyxl


def convert_pdf_to_excel(input_pdf_file, output_excel_file):
    # Read the PDF and get the "No de commande" and "Nom de client"
    no_commande, nom_client = read_pdf_and_get_no_commande(input_pdf_file)

    if not no_commande or not nom_client:
        print("No commande or Nom de client not found in the PDF.")
        return
    a = 0
    # Create Excel Workbook
    wb = Workbook()
    # Liste pour stocker les DataFrames des feuilles valides
    valid_dfs = []
    # Initialize an empty list to store errors
    errors = []
    error_message = []
    nombre_total_de_lignes = 0
    # Read PDF File and Extract Tables
    # Read PDF File and Extract Tables
    with pdfplumber.open(input_pdf_file) as pdf:

        for i, page in enumerate(pdf.pages, start=1):
            table = page.extract_table()
            if table:
                header = table[0]  # Get the header row
                if "Article" in header:
                    combined_data = table[1:]  # Exclude header row
                    df = pd.DataFrame(combined_data, columns=header)
                    # Convert "Article" column to string type
                    df["Article"] = df["Article"].astype(str)

                    # Create a new column "UV" to store the corresponding UV value
                    df["UV"] = None

                    # Match "Article" column with "Code_Barre" in the bondedecommande table
                    articles_without_match = []
                    error_article = []
                    for index, row in df.iterrows():
                        article = row["Article"]
                        code_barre_id = bondedecommandeeee.objects.filter(Code_Barre=article).values_list('id',
                                                                                                         flat=True).first()

                        if code_barre_id is not None:
                            code_article_magasin_id = bondedecommandeeee.objects.filter(id=code_barre_id).values_list(
                                'Code_Article_magasin', flat=True).first()
                            uv = bondedecommandee.objects.filter(Code_Barre=article).values_list('UV',
                                                                                                 flat=True).first()

                            df.at[index, "Article"] = code_article_magasin_id

                            print(
                                f"Article: {article} | Code_Barre ID: {code_barre_id} | Code_Article_magasin ID: {code_article_magasin_id}")
                        else:
                            print(article)

                            error_message.append("Code_Barre not found in database")
                            error_message.append(article)
                            raise ValueError(error_message)

                    print("Articles without a match in the database:")
                    print(articles_without_match)

                    valid_dfs.append(df)

                else:
                    print(f"Table {i} does not contain the required columns.")

        # Après avoir traité tous les PDF, imprimez le nombre total de lignes
        # Ajouter le DataFrame à la liste des DataFrames valides

        #        print(df, "sanaae")
        # Save the DataFrame to the Excel sheet
        sheet_name = f"Table_{i}"
        ws = wb.create_sheet(title=sheet_name)
        for r_idx, row in enumerate(dataframe_to_rows(df, index=False), start=1):
            for c_idx, value in enumerate(row, start=1):
                ws.cell(row=r_idx, column=c_idx, value=value)

    # ...

    # Save the Excel Workbook
    wb.save(output_excel_file)
    dfs_to_concat = []  # List to store valid DataFrames for concatenation

    # Read the Excel file again to get the valid sheet names
    valid_sheet_names = [f"Table_{j}" for j in range(1, len(valid_dfs) + 1)]

    for sheet_name in valid_sheet_names:
        try:
            df = pd.read_excel(output_excel_file, sheet_name=sheet_name)

            # Vérifier si la colonne "Article" existe
            if "Article" in df.columns:
                dfs_to_concat.append(df)
            else:
                print(f"La feuille {sheet_name} ne contient pas la colonne 'Article' requise.")
        except Exception as e:
            print(f"Erreur lors de la lecture de la feuille {sheet_name} : {e}")

        # ...

        # Rest of your code for concatenating DataFrames and saving the final DataFrame

        if valid_dfs:
            combined_df = pd.concat(valid_dfs, ignore_index=True)
            combined_df = combined_df[~combined_df["Article"].astype(str).str.contains("Article")]
            print("sanaa", combined_df)
            # Assuming you have the combined_df DataFrame somewhere in your code
            # generate_report(combined_df)

            # Enregistrer le DataFrame combiné dans le fichier Excel
            combined_df.to_excel(output_excel_file, sheet_name="Combined_Table", index=False)
            print(combined_df.shape[0], "nombre de ligne de table article de ce pdf selectionner")
        else:
            print("Aucune feuille valide trouvée pour la concaténation.")

    print("All tables combined and saved to Excel.")
    print(f"No commande: {no_commande}")
    print(f"nom de client: {nom_client}")

    # Variable to store the id of the corresponding client Quanten\nUVC
    matching_client_id = None

    clients = Client.objects.all()

    # Loop through the records and print each column's value
    for client in clients:
        # Remove spaces from the "Nom_Client" field of the current client object
        nom_client_stripped = client.Nom_Client.replace(" ", "")

        # Compare the stripped Nom_Client from the PDF with the stripped Nom_Client from the database
        if nom_client_stripped.lower() == nom_client.lower():
            print("Comparison Result: True (Match)")
            # Save the id of the corresponding client
            matching_client_id = client.id

            # Display the information of the corresponding client
            print(f"ID du client correspondant: {matching_client_id}")
            print(f"Groupe_Vendeur: {client.Groupe_Vendeur}")
            print(f"Designation_Vendeur: {client.Designation_Vendeur}")
            print(f"Conditions_paiement: {client.Conditions_paiement}")
            print(f"Agence_commerciale: {client.Agence_commerciale}")
            print(f"Nom: {client.Nom_Client}")
            print(f"DIVISON: {client.DIVISON}")
            print("\n")

            # Exit the loop as we have found a match
            break

    # Check if a matching client has been found
    if matching_client_id:
        print(f"Le client correspondant a été trouvé.")

        # Add the "Groupe_Vendeur", "Designation_Vendeur", and "Conditions_paiement" columns to the "Combined_Table" DataFrame
        import datetime

        combined_df.insert(0, "Type doc vte", "ZNCS")
        combined_df.insert(1, "Organis.commerciale", "V500")
        combined_df.insert(2, "Canal_distribution", "GS")
        combined_df.insert(3, "Secteur_activite", "ZN")
        combined_df.insert(4, "Agence", client.Agence_commerciale)
        combined_df.insert(5, "Groupe de vendeurs", client.Groupe_Vendeur)

        import datetime
        today_date = datetime.date.today().strftime("%Y.%m.%d")
        # combined_df.insert(5, "Nom", client.Nom_Client)
        current_date = datetime.date.today().strftime("%Y-%m-%d")
        combined_df.insert(6, "Donneur", client.Code_Client)
        combined_df.insert(7, " Rece", client.Code_Client)
        combined_df.insert(8, "N", no_commande)
        combined_df.insert(9, "Date", today_date)
        combined_df.insert(11, "Num", "")
        combined_df.insert(13, "Itinéraire", "")
        combined_df.insert(14, "Division", client.DIVISON)
        combined_df.insert(15, "Poids", "")
        combined_df.insert(16, "TXLINE_02", "")
        combined_df.insert(17, "N équipement", "")
        # combined_df.insert(9, " Article", article)
        # combined_df.insert(10, "unité de vente", uv)
        # combined_df.insert(11, "Type poste", "TAN")
        # combined_df.insert(12, "DIVISION", client.DIVISON)

        # Save the updated DataFrame to the Excel file with the new columns
        combined_df.to_excel(output_excel_file, sheet_name="Combined_Table", index=False)

        # ...
        # combined_df["Type poste"] ="TAN"
        # combined_df["DIVISON"] = client.DIVISON

        # combined_df["Désignation"] = client.Designation_Vendeur
        # combined_df["Conditions paiement"] = client.Conditions_paiement

        # Remove the columns VL and Noligne from the DataFrame
        combined_df.drop(columns=["VL", "Noligne", "TypeU.C.", "UVC/UC", "UV", "Libellearticle"], inplace=True)

        # Remove columns that end with the word "speciale" from the DataFrame
        columns_to_drop = [col for col in combined_df.columns if col.lower().endswith("speciale")]
        combined_df.drop(columns=columns_to_drop, inplace=True)

        # Remove columns that end with the word "UC" from the DataFrame
        columns_to_drop_UC = [col for col in combined_df.columns if col.lower().endswith("UC")]
        combined_df.drop(columns=columns_to_drop_UC, inplace=True)
        # Remove columns that end with the word "UC" from the DataFrame
        # columns_to_drop_libelle = [col for col in combined_df.columns if col.lower().endswith("Libellearticle")]
        # combined_df.drop(columns=columns_to_drop_libelle, inplace=True)
        first_column_to_drop = next((col for col in combined_df.columns if col.lower().startswith("quanten")), None)
        if first_column_to_drop:
            combined_df.drop(columns=first_column_to_drop, inplace=True)
            # Renommer la colonne "quanten" si elle existe
        column_to_replace = next((col for col in combined_df.columns if col.lower().startswith("quanten")), None)
        if column_to_replace:
            combined_df.rename(columns={column_to_replace: "quantité"}, inplace=True)

        # Déplacer la colonne "quantité d'ordre" après la colonne "Article"
        if "Num" in combined_df.columns and "quantité" in combined_df.columns:
            article_index = combined_df.columns.get_loc("Num")
            quantity_order_index = combined_df.columns.get_loc("quantité")
            if quantity_order_index != article_index + 1:
                combined_df = combined_df[combined_df.columns[:article_index + 1].tolist() +
                                          combined_df.columns[quantity_order_index:quantity_order_index + 1].tolist() +
                                          combined_df.columns[article_index + 1:quantity_order_index].tolist() +
                                          combined_df.columns[quantity_order_index + 1:].tolist()]
        first_column_to_drop_article = next((col for col in combined_df.columns if col.lower().startswith("Num")), None)
        if first_column_to_drop_article:
            combined_df.drop(columns=first_column_to_drop_article, inplace=True)

            # column_to_replace_libelle = next((col for col in combined_df.columns if col.lower().startswith("Libelle article")), None)
            # if column_to_replace_libelle:
            #     combined_df.rename(columns={column_to_replace_libelle: "Désignation article"}, inplace=True)
        # Save the updated DataFrame to the Excel file with the new columns
        # Renommer la colonne "Article" en "Articlee"
        if "Article" in combined_df.columns:
            combined_df.rename(columns={"Article": "Articlee"}, inplace=True)
        combined_df.to_excel(output_excel_file, sheet_name="Combined_Table", index=False)
    else:
        print(f"Erreur:'{nom_client}'")
        # error_message_client = f"Le client n'a pas été trouvé dans la base de données pour '{nom_client}'"
        error_message.append("Nom client not found")
        error_message.append(nom_client)
        raise ValueError(error_message)

        # Call the generate_error_excel function to create an error Excel file
    # Load the combined DataFrame from the Excel file
    combined_df = pd.read_excel(output_excel_file, sheet_name="Combined_Table")

    # Check if the "quantite" column exists
    if "quantité" in combined_df.columns and "Articlee" in combined_df.columns:
        for index, row in combined_df.iterrows():
            article = row["Articlee"]
            quantite = row["quantité"]

            # Query the database to retrieve the ID based on the "Article"
            code_barre_id = bondedecommandeeee.objects.filter(Code_Article_magasin=article).values_list('Colisage',
                                                                                                       flat=True).first()
            # Divide the values in the "quantite" column by 2
            #combined_df.at[index, "quantité"] = quantite / code_barre_id
            combined_df.at[index, "quantité"] = int(quantite / code_barre_id)
            # Save the updated DataFrame back to the Excel file
    combined_df.to_excel(output_excel_file, sheet_name="Combined_Table", index=False)

    print("Quantite column updated and saved to Excel.")

    return output_excel_file

    #    else:
    #
    #    print(f"Aucun client correspondant n'a été trouvé.")


def read_pdf_and_get_no_commande(pdf_file):
    # Read PDF File and Extract Text using OCR (Tesseract)
    with pdfplumber.open(pdf_file) as pdf:
        first_page = pdf.pages[0]
        text = first_page.extract_text()

        # Split the text into lines
        lines = text.splitlines()
        # Print the extracted text for inspection
        print("Extracted Text:")
        print(text)
        # Check if there are at least 5 lines in the text
        if len(lines) >= 5:
            # Get the content of the first column from the fifth line as "No de commande"
            no_commande = lines[2].split()[0]

            # Get the content of the first column from the fifth line as "Nom de client"
            nom_client = lines[4].split()[0]

            return no_commande, nom_client

        else:
            print("No commande or Nom de client not found in the PDF.")
            return None, None

def read_pdf_and_get_no_commande_client(pdf_file):
    # Read PDF File and Extract Text using OCR (Tesseract)
    with pdfplumber.open(pdf_file) as pdf:
        first_page = pdf.pages[0]
        text = first_page.extract_text()

        # Split the text into lines
        lines = text.splitlines()
        # Print the extracted text for inspection
        print("Extracted Text:")
        print(text)
        # Check if there are at least 5 lines in the text
        if len(lines) >= 5:
            # Get the content of the first column from the fifth line as "No de commande"
            no_commande = lines[6].split()[0]

            # Get the content of the first column from the fifth line as "Nom de client"
            nom_client = lines[8].split()[0]

            return no_commande, nom_client

        else:
            print("No commande or Nom de client not found in the PDF.")
            return None, None

import json
from django.http import JsonResponse
import os
from openpyxl import load_workbook, Workbook
from django.http import HttpResponse
from django.shortcuts import render
from django.shortcuts import render
from django.http import HttpResponseServerError
from .models import Report  # Import the Report model
import os
from django.shortcuts import render

from django.shortcuts import render
from django.http import JsonResponse
from .models import Report
import os


def report_list(request):
    reports = Report.objects.all()
    return render(request, 'Myapp/report_list.html', {'reports': reports})


def download_report(request, report_id):
    report = Report.objects.get(id=report_id)
    response = HttpResponse(report.excel_content,
                            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{report.excel_filename}"'
    return response


from django.http import JsonResponse
from django.shortcuts import render

# ... (other imports)

from django.http import JsonResponse
from django.shortcuts import render

# ... (other imports)

from django.http import JsonResponse
from django.shortcuts import render

# ... (other imports)

from django.http import JsonResponse, FileResponse
import os
import datetime
from .models import Report

from django.shortcuts import render
from openpyxl import load_workbook


def comparaisonExcel(request):
    if request.method == 'POST':
        excel_file = request.FILES.get('excel_file')
        sap_excel_file = request.FILES.get('sap_excel_file')

        if excel_file and sap_excel_file:
            try:
                excel_data = extract_data_from_excel(excel_file)
                sap_excel_data = extract_data_from_excel(sap_excel_file)

                comparison_results = []

                for row in excel_data:
                    n_value = row['N']
                    exists_in_sap = any(sap_row['Nº commande d\'achat'] == n_value for sap_row in sap_excel_data)
                    comparison_results.append({'N': n_value, 'Exists in SAP': exists_in_sap})

                return render(request, 'Myapp/comparaisonExcel.html', {'results': comparison_results})

            except Exception as e:
                error_message = str(e)
                return render(request, 'error.html', {'error_message': error_message})

    return render(request, 'Myapp/comparaisonExcel.html')


def extract_data_from_excel(file):
    workbook = load_workbook(file)
    sheet = workbook.active
    headers = [cell.value for cell in sheet[1]]

    data = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        data.append(dict(zip(headers, row)))

    return data


def report_list(request):
    reports = Report.objects.all()
    return render(request, 'Myapp/report_list.html', {'reports': reports})


def download_report(request, report_id):
    report = Report.objects.get(id=report_id)
    response = HttpResponse(report.excel_content,
                            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{report.excel_filename}"'
    return response


import os
import zipfile
from django.http import HttpResponse
from django.shortcuts import render
from django.conf import settings


def download_zip(request, zip_file_path):
    with open(zip_file_path, 'rb') as zip_file:
        response = HttpResponse(zip_file.read(), content_type='application/zip')
        response['Content-Disposition'] = f'attachment; filename={os.path.basename(zip_file_path)}'
        return response


import os
import openpyxl
from django.http import HttpResponse
from django.shortcuts import render
from django.conf import settings

import os
import openpyxl
from django.http import HttpResponse
from django.shortcuts import render
from django.conf import settings

import os
import openpyxl
from django.http import HttpResponse
from django.shortcuts import render
from django.conf import settings
import zipfile
from django.http import HttpResponse
import io
import shutil


def index(request):
    if request.method == "POST":
        pdf_files = request.FILES.getlist("pdf_file")

        if pdf_files:
            media_dir = os.path.join(settings.MEDIA_ROOT, "uploads")
            if not os.path.exists(media_dir):
                os.makedirs(media_dir)

            excel_files = []
            errors = {}
            success_files = []
            for pdf_file in pdf_files:
                pdf_file_path = os.path.join(media_dir, pdf_file.name)

                with open(pdf_file_path, 'wb') as destination:
                    for chunk in pdf_file.chunks():
                        destination.write(chunk)

                try:
                    excel_file_path = os.path.join(media_dir, f"{os.path.splitext(pdf_file.name)[0]}.xlsx")
                    excel_file_path = convert_pdf_to_excel(pdf_file_path, excel_file_path)
                    excel_files.append(excel_file_path)
                    success_files.append(pdf_file.name)
                except ValueError as e:
                    error_message = str(e)
                    errors[pdf_file.name] = error_message

            # Si tous les fichiers PDF ont échoué
            if not success_files and errors:
                # Créer un fichier Excel pour stocker les erreurs
                error_excel = openpyxl.Workbook()
                error_sheet = error_excel.active
                # Add headers to the new sheet
                error_sheet.append(['pdf_name', 'Remarque Client/Article'])
                # Écrire les erreurs dans le fichier Excel
                # Écrire les erreurs dans le fichier Excel
                for i, (file_name, error_message) in enumerate(errors.items(), start=1):
                    error_sheet.cell(row=i + 1, column=1, value=file_name)
                    error_sheet.cell(row=i + 1, column=2, value=error_message)

                # Sauvegarder le fichier Excel
                error_report_path = os.path.join(media_dir, 'error_report.xlsx')
                error_excel.save(error_report_path)

                # Vous pouvez maintenant retourner ce fichier Excel comme réponse
                with open(error_report_path, 'rb') as excel_file:
                    response = HttpResponse(excel_file.read(),
                                            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                    response['Content-Disposition'] = 'attachment; filename=error_report.xlsx'
                    return response

            # Combine Excel files into a single Excel file
            combined_excel_file_path = os.path.join(media_dir, "combined_excel.xlsx")
            combine_excel_files(excel_files, combined_excel_file_path)

            if errors and success_files:
                # Open the combined Excel file
                combined_excel = openpyxl.load_workbook(combined_excel_file_path)

                # Remove the empty sheet if it exists (often named 'Sheet')
                if 'Sheet' in combined_excel.sheetnames:
                    sheet_to_remove = combined_excel['Sheet']
                    combined_excel.remove(sheet_to_remove)

                # Create a new sheet in 'combined_excel.xlsx' to copy data from 'error_report.xlsx'
                new_sheet = combined_excel.create_sheet(title='error_report_data')

                # Add headers to the new sheet
                new_sheet.append(['pdf_name', 'Remarque Client/Article'])

                # Open the 'error_report.xlsx' file or create it if it doesn't exist
                error_report_path = os.path.join(media_dir, 'error_report.xlsx')
                if os.path.exists(error_report_path):
                    error_report = openpyxl.load_workbook(error_report_path)
                else:
                    error_report = openpyxl.Workbook()

                # Get or create the 'error_report_data' sheet
                if 'error_report_data' not in error_report.sheetnames:
                    error_report_sheet = error_report.create_sheet(title='error_report_data')
                else:
                    error_report_sheet = error_report['error_report_data']

                # Add headers to the new sheet if it's newly created
                if error_report_sheet.max_row == 1:
                    error_report_sheet.append(['pdf_name', 'Remarque Client/Article'])

                # Add errors from the PDF files that failed to the 'error_report_data' sheet
                for pdf_file, error_message in errors.items():
                    error_report_sheet.append([pdf_file, error_message])

                # Save the 'error_report.xlsx' file
                error_report.save(error_report_path)

                # Add errors from the PDF files that failed to the 'error_report_data' sheet in the combined Excel
                for pdf_file, error_message in errors.items():
                    new_sheet.append([pdf_file, error_message])

                # Save the combined Excel file (without the empty sheet)
                combined_excel.save(combined_excel_file_path)

                # Generate the content of the text file from the combined Excel file
                text_content = generate_text_content(combined_excel_file_path)

                # Create a ZIP file
                zip_buffer = io.BytesIO()
                with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zipf:
                    # Add the combined Excel file to the ZIP
                    zipf.write(combined_excel_file_path, arcname='combined_excel_with_errors.xlsx')

                    # Add the 'error_report.xlsx' file to the ZIP
                    zipf.write(error_report_path, arcname='error_report.xlsx')

                    # Add the text file to the ZIP
                    zipf.writestr('combined_blocnote.txt', text_content)

                # Prepare the HTTP response for the ZIP file
                response = HttpResponse(zip_buffer.getvalue(), content_type='application/zip')
                response['Content-Disposition'] = 'attachment; filename=combined_files.zip'

                return response


            elif success_files:

                if remove_empty_sheet(combined_excel_file_path):
                    # Générer le contenu du fichier texte à partir du fichier Excel combiné
                    text_content = generate_text_content(combined_excel_file_path)

                    # Télécharger le fichier Excel combiné
                    excel_response = download_excel(request, combined_excel_file_path)

                    # Créer un fichier ZIP
                    zip_buffer = io.BytesIO()
                    with zipfile.ZipFile(zip_buffer, 'w') as zipf:
                        # Ajouter le fichier Excel au ZIP
                        zipf.writestr('combined_excel.xlsx', excel_response.content)
                        # Ajouter le fichier texte au ZIP
                        zipf.writestr('combined_blocnote.txt', text_content)

                    # Définir les en-têtes appropriés pour le téléchargement du fichier ZIP
                    response = HttpResponse(zip_buffer.getvalue(), content_type='application/zip')
                    response['Content-Disposition'] = 'attachment; filename=combined_files.zip'

                    return response

    return render(request, 'Myapp/index.html')


import openpyxl


def generate_text_content(excel_file_path):
    # Charger le fichier Excel en utilisant pandas
    df = pd.read_excel(excel_file_path, engine='openpyxl')
    # Convertir le DataFrame en une chaîne de texte au format CSV
    #text_content = df.to_csv(index=False, header=True, sep='\t', na_rep='')
    text_content = df.to_csv(index=False, header=False, sep='\t', na_rep='')

    return text_content

def convert_pdf_to_excel_itkaneaswak(input_pdf_file, output_excel_file):
    # Read the PDF and get the "No de commande" and "Nom de client"
    no_commande, nom_client = read_pdf_and_get_no_commande_client(input_pdf_file)

    if not no_commande or not nom_client:
        print("No commande or Nom de client not found in the PDF.")
        return
    a = 0
    # Create Excel Workbook
    wb = Workbook()
    # Liste pour stocker les DataFrames des feuilles valides
    valid_dfs = []
    # Initialize an empty list to store errors
    errors = []
    error_message = []
    nombre_total_de_lignes = 0
    # Read PDF File and Extract Tables
    # Read PDF File and Extract Tables
    with pdfplumber.open(input_pdf_file) as pdf:

        for i, page in enumerate(pdf.pages, start=1):
            table = page.extract_table()
            if table:
                header = table[0]  # Get the header row
                if "Article" in header:
                    combined_data = table[1:]  # Exclude header row
                    df = pd.DataFrame(combined_data, columns=header)
                    # Convert "Article" column to string type
                    df["Article"] = df["Article"].astype(str)

                    # Create a new column "UV" to store the corresponding UV value
                    df["UV"] = None

                    # Match "Article" column with "Code_Barre" in the bondedecommande table
                    articles_without_match = []
                    error_article = []
                    for index, row in df.iterrows():
                        article = row["Article"]
                        code_barre_id = bondedecommandeeee.objects.filter(Code_Barre=article).values_list('id',
                                                                                                         flat=True).first()

                        if code_barre_id is not None:
                            code_article_magasin_id = bondedecommandeeee.objects.filter(id=code_barre_id).values_list(
                                'Code_Article_magasin', flat=True).first()
                            uv = bondedecommandee.objects.filter(Code_Barre=article).values_list('UV',
                                                                                                 flat=True).first()

                            df.at[index, "Article"] = code_article_magasin_id

                            print(
                                f"Article: {article} | Code_Barre ID: {code_barre_id} | Code_Article_magasin ID: {code_article_magasin_id}")
                        else:
                            print(article)

                            error_message.append("Code_Barre not found in database")
                            error_message.append(article)
                            raise ValueError(error_message)

                    print("Articles without a match in the database:")
                    print(articles_without_match)

                    valid_dfs.append(df)

                else:
                    print(f"Table {i} does not contain the required columns.")

        # Après avoir traité tous les PDF, imprimez le nombre total de lignes
        # Ajouter le DataFrame à la liste des DataFrames valides

        #        print(df, "sanaae")
        # Save the DataFrame to the Excel sheet
        sheet_name = f"Table_{i}"
        ws = wb.create_sheet(title=sheet_name)
        for r_idx, row in enumerate(dataframe_to_rows(df, index=False), start=1):
            for c_idx, value in enumerate(row, start=1):
                ws.cell(row=r_idx, column=c_idx, value=value)

    # ...

    # Save the Excel Workbook
    wb.save(output_excel_file)
    dfs_to_concat = []  # List to store valid DataFrames for concatenation

    # Read the Excel file again to get the valid sheet names
    valid_sheet_names = [f"Table_{j}" for j in range(1, len(valid_dfs) + 1)]

    for sheet_name in valid_sheet_names:
        try:
            df = pd.read_excel(output_excel_file, sheet_name=sheet_name)

            # Vérifier si la colonne "Article" existe
            if "Article" in df.columns:
                dfs_to_concat.append(df)
            else:
                print(f"La feuille {sheet_name} ne contient pas la colonne 'Article' requise.")
        except Exception as e:
            print(f"Erreur lors de la lecture de la feuille {sheet_name} : {e}")

        # ...

        # Rest of your code for concatenating DataFrames and saving the final DataFrame

        if valid_dfs:
            combined_df = pd.concat(valid_dfs, ignore_index=True)
            combined_df = combined_df[~combined_df["Article"].astype(str).str.contains("Article")]
            print("sanaa", combined_df)
            # Assuming you have the combined_df DataFrame somewhere in your code
            # generate_report(combined_df)

            # Enregistrer le DataFrame combiné dans le fichier Excel
            combined_df.to_excel(output_excel_file, sheet_name="Combined_Table", index=False)
            print(combined_df.shape[0], "nombre de ligne de table article de ce pdf selectionner")
        else:
            print("Aucune feuille valide trouvée pour la concaténation.")

    print("All tables combined and saved to Excel.")
    print(f"No commande: {no_commande}")
    print(f"nom de client: {nom_client}")

    # Variable to store the id of the corresponding client Quanten\nUVC
    matching_client_id = None

    clients = Client.objects.all()

    # Loop through the records and print each column's value
    for client in clients:
        # Remove spaces from the "Nom_Client" field of the current client object
        nom_client_stripped = client.Nom_Client.replace(" ", "")

        # Compare the stripped Nom_Client from the PDF with the stripped Nom_Client from the database
        if nom_client_stripped.lower() == nom_client.lower():
            print("Comparison Result: True (Match)")
            # Save the id of the corresponding client
            matching_client_id = client.id

            # Display the information of the corresponding client
            print(f"ID du client correspondant: {matching_client_id}")
            print(f"Groupe_Vendeur: {client.Groupe_Vendeur}")
            print(f"Designation_Vendeur: {client.Designation_Vendeur}")
            print(f"Conditions_paiement: {client.Conditions_paiement}")
            print(f"Agence_commerciale: {client.Agence_commerciale}")
            print(f"Nom: {client.Nom_Client}")
            print(f"DIVISON: {client.DIVISON}")
            print("\n")

            # Exit the loop as we have found a match
            break

    # Check if a matching client has been found
    if matching_client_id:
        print(f"Le client correspondant a été trouvé.")

        # Add the "Groupe_Vendeur", "Designation_Vendeur", and "Conditions_paiement" columns to the "Combined_Table" DataFrame
        import datetime

        combined_df.insert(0, "Type doc vte", "ZNCS")
        combined_df.insert(1, "Organis.commerciale", "V500")
        combined_df.insert(2, "Canal_distribution", "GS")
        combined_df.insert(3, "Secteur_activite", "ZN")
        combined_df.insert(4, "Agence", client.Agence_commerciale)
        combined_df.insert(5, "Groupe de vendeurs", client.Groupe_Vendeur)

        import datetime
        today_date = datetime.date.today().strftime("%Y.%m.%d")
        # combined_df.insert(5, "Nom", client.Nom_Client)
        current_date = datetime.date.today().strftime("%Y-%m-%d")
        combined_df.insert(6, "Donneur", client.Code_Client)
        combined_df.insert(7, " Rece", client.Code_Client)
        combined_df.insert(8, "N", no_commande)
        combined_df.insert(9, "Date", today_date)
        combined_df.insert(11, "Num", "")
        combined_df.insert(13, "Itinéraire", "")
        combined_df.insert(14, "Division", client.DIVISON)
        combined_df.insert(15, "Poids", "")
        combined_df.insert(16, "TXLINE_02", "")
        combined_df.insert(17, "N équipement", "")
        # combined_df.insert(9, " Article", article)
        # combined_df.insert(10, "unité de vente", uv)
        # combined_df.insert(11, "Type poste", "TAN")
        # combined_df.insert(12, "DIVISION", client.DIVISON)

        # Save the updated DataFrame to the Excel file with the new columns
        combined_df.to_excel(output_excel_file, sheet_name="Combined_Table", index=False)

        # ...
        # combined_df["Type poste"] ="TAN"
        # combined_df["DIVISON"] = client.DIVISON

        # combined_df["Désignation"] = client.Designation_Vendeur
        # combined_df["Conditions paiement"] = client.Conditions_paiement

        # Remove the columns VL and Noligne from the DataFrame
        combined_df.drop(columns=["VL", "Noligne", "UVC/UC", "UV", "Libellearticle","Qteenstock","Type\nU.C."], inplace=True)

        # Remove columns that end with the word "speciale" from the DataFrame
        columns_to_drop = [col for col in combined_df.columns if col.lower().endswith("speciale")]
        combined_df.drop(columns=columns_to_drop, inplace=True)

        # Remove columns that end with the word "UC" from the DataFrame
        columns_to_drop_UC = [col for col in combined_df.columns if col.lower().endswith("UC")]
        combined_df.drop(columns=columns_to_drop_UC, inplace=True)
        # Remove columns that end with the word "UC" from the DataFrame
        # columns_to_drop_libelle = [col for col in combined_df.columns if col.lower().endswith("Libellearticle")]
        # combined_df.drop(columns=columns_to_drop_libelle, inplace=True)


        # Déplacer la colonne "quantité d'ordre" après la colonne "Article"
        if "Num" in combined_df.columns and "Quanten\nUC" in combined_df.columns:
            article_index = combined_df.columns.get_loc("Num")
            quantity_order_index = combined_df.columns.get_loc("Quanten\nUC")
            if quantity_order_index != article_index + 1:
                combined_df = combined_df[combined_df.columns[:article_index + 1].tolist() +
                                          combined_df.columns[quantity_order_index:quantity_order_index + 1].tolist() +
                                          combined_df.columns[article_index + 1:quantity_order_index].tolist() +
                                          combined_df.columns[quantity_order_index + 1:].tolist()]
        first_column_to_drop_article = next((col for col in combined_df.columns if col.lower().startswith("Num")), None)
        if first_column_to_drop_article:
            combined_df.drop(columns=first_column_to_drop_article, inplace=True)

            # column_to_replace_libelle = next((col for col in combined_df.columns if col.lower().startswith("Libelle article")), None)
            # if column_to_replace_libelle:
            #     combined_df.rename(columns={column_to_replace_libelle: "Désignation article"}, inplace=True)
        # Save the updated DataFrame to the Excel file with the new columns
        # Renommer la colonne "Article" en "Articlee"
        if "Article" in combined_df.columns:
            combined_df.rename(columns={"Article": "Articlee"}, inplace=True)
        combined_df.to_excel(output_excel_file, sheet_name="Combined_Table", index=False)
    else:
        print(f"Erreur:'{nom_client}'")
        # error_message_client = f"Le client n'a pas été trouvé dans la base de données pour '{nom_client}'"
        error_message.append("Nom client not found")
        error_message.append(nom_client)
        raise ValueError(error_message)

        # Call the generate_error_excel function to create an error Excel file
    # Load the combined DataFrame from the Excel file
    combined_df = pd.read_excel(output_excel_file, sheet_name="Combined_Table")

    # Check if the "quantite" column exists
    if "Quanten\nUC" in combined_df.columns and "Articlee" in combined_df.columns:
        for index, row in combined_df.iterrows():
            article = row["Articlee"]
            quantite = row["Quanten\nUC"]

            # Query the database to retrieve the ID based on the "Article"
            code_barre_id = bondedecommandeeee.objects.filter(Code_Article_magasin=article).values_list('Colisage',
                                                                                                       flat=True).first()
            # Divide the values in the "quantite" column by 2
            #combined_df.at[index, "quantité"] = quantite / code_barre_id
            combined_df.at[index, "Quanten\nUC"] = quantite

            # Save the updated DataFrame back to the Excel file
    combined_df.to_excel(output_excel_file, sheet_name="Combined_Table", index=False)

    print("Quantite column updated and saved to Excel.")

    return output_excel_file

    #    else:
    #
    #    print(f"Aucun client correspondant n'a été trouvé.")
def indexxx(request):
    if request.method == "POST":
        pdf_files = request.FILES.getlist("pdf_file")

        if pdf_files:
            media_dir = os.path.join(settings.MEDIA_ROOT, "uploads")
            if not os.path.exists(media_dir):
                os.makedirs(media_dir)

            excel_files = []
            errors = {}
            success_files = []
            for pdf_file in pdf_files:
                pdf_file_path = os.path.join(media_dir, pdf_file.name)

                with open(pdf_file_path, 'wb') as destination:
                    for chunk in pdf_file.chunks():
                        destination.write(chunk)

                try:
                    excel_file_path = os.path.join(media_dir, f"{os.path.splitext(pdf_file.name)[0]}.xlsx")
                    excel_file_path = convert_pdf_to_excel_itkaneaswak(pdf_file_path, excel_file_path)
                    excel_files.append(excel_file_path)
                    success_files.append(pdf_file.name)
                except ValueError as e:
                    error_message = str(e)
                    errors[pdf_file.name] = error_message

            # Si tous les fichiers PDF ont échoué
            if not success_files and errors:
                # Créer un fichier Excel pour stocker les erreurs
                error_excel = openpyxl.Workbook()
                error_sheet = error_excel.active
                # Add headers to the new sheet
                error_sheet.append(['pdf_name', 'Remarque Client/Article'])
                # Écrire les erreurs dans le fichier Excel
                # Écrire les erreurs dans le fichier Excel
                for i, (file_name, error_message) in enumerate(errors.items(), start=1):
                    error_sheet.cell(row=i + 1, column=1, value=file_name)
                    error_sheet.cell(row=i + 1, column=2, value=error_message)

                # Sauvegarder le fichier Excel
                error_report_path = os.path.join(media_dir, 'error_report.xlsx')
                error_excel.save(error_report_path)

                # Vous pouvez maintenant retourner ce fichier Excel comme réponse
                with open(error_report_path, 'rb') as excel_file:
                    response = HttpResponse(excel_file.read(),
                                            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                    response['Content-Disposition'] = 'attachment; filename=error_report.xlsx'
                    return response

            # Combine Excel files into a single Excel file
            combined_excel_file_path = os.path.join(media_dir, "combined_excel.xlsx")
            combine_excel_files(excel_files, combined_excel_file_path)

            if errors and success_files:
                # Open the combined Excel file
                combined_excel = openpyxl.load_workbook(combined_excel_file_path)

                # Remove the empty sheet if it exists (often named 'Sheet')
                if 'Sheet' in combined_excel.sheetnames:
                    sheet_to_remove = combined_excel['Sheet']
                    combined_excel.remove(sheet_to_remove)

                # Create a new sheet in 'combined_excel.xlsx' to copy data from 'error_report.xlsx'
                new_sheet = combined_excel.create_sheet(title='error_report_data')

                # Add headers to the new sheet
                new_sheet.append(['pdf_name', 'Remarque Client/Article'])

                # Open the 'error_report.xlsx' file or create it if it doesn't exist
                error_report_path = os.path.join(media_dir, 'error_report.xlsx')
                if os.path.exists(error_report_path):
                    error_report = openpyxl.load_workbook(error_report_path)
                else:
                    error_report = openpyxl.Workbook()

                # Get or create the 'error_report_data' sheet
                if 'error_report_data' not in error_report.sheetnames:
                    error_report_sheet = error_report.create_sheet(title='error_report_data')
                else:
                    error_report_sheet = error_report['error_report_data']

                # Add headers to the new sheet if it's newly created
                if error_report_sheet.max_row == 1:
                    error_report_sheet.append(['pdf_name', 'Remarque Client/Article'])

                # Add errors from the PDF files that failed to the 'error_report_data' sheet
                for pdf_file, error_message in errors.items():
                    error_report_sheet.append([pdf_file, error_message])

                # Save the 'error_report.xlsx' file
                error_report.save(error_report_path)

                # Add errors from the PDF files that failed to the 'error_report_data' sheet in the combined Excel
                for pdf_file, error_message in errors.items():
                    new_sheet.append([pdf_file, error_message])

                # Save the combined Excel file (without the empty sheet)
                combined_excel.save(combined_excel_file_path)

                # Generate the content of the text file from the combined Excel file
                text_content = generate_text_content(combined_excel_file_path)

                # Create a ZIP file
                zip_buffer = io.BytesIO()
                with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zipf:
                    # Add the combined Excel file to the ZIP
                    zipf.write(combined_excel_file_path, arcname='combined_excel_with_errors.xlsx')

                    # Add the 'error_report.xlsx' file to the ZIP
                    zipf.write(error_report_path, arcname='error_report.xlsx')

                    # Add the text file to the ZIP
                    zipf.writestr('combined_blocnote.txt', text_content)

                # Prepare the HTTP response for the ZIP file
                response = HttpResponse(zip_buffer.getvalue(), content_type='application/zip')
                response['Content-Disposition'] = 'attachment; filename=combined_files.zip'

                return response


            elif success_files:

                if remove_empty_sheet(combined_excel_file_path):
                    # Générer le contenu du fichier texte à partir du fichier Excel combiné
                    text_content = generate_text_content(combined_excel_file_path)

                    # Télécharger le fichier Excel combiné
                    excel_response = download_excel(request, combined_excel_file_path)

                    # Créer un fichier ZIP
                    zip_buffer = io.BytesIO()
                    with zipfile.ZipFile(zip_buffer, 'w') as zipf:
                        # Ajouter le fichier Excel au ZIP
                        zipf.writestr('combined_excel.xlsx', excel_response.content)
                        # Ajouter le fichier texte au ZIP
                        zipf.writestr('combined_blocnote.txt', text_content)

                    # Définir les en-têtes appropriés pour le téléchargement du fichier ZIP
                    response = HttpResponse(zip_buffer.getvalue(), content_type='application/zip')
                    response['Content-Disposition'] = 'attachment; filename=combined_files.zip'

                    return response

    return render(request, 'Myapp/indexxx.html')


import openpyxl
def index4(request):
    if request.method == "POST":
        pdf_files = request.FILES.getlist("pdf_file")

        if pdf_files:
            media_dir = os.path.join(settings.MEDIA_ROOT, "uploads")
            if not os.path.exists(media_dir):
                os.makedirs(media_dir)

            excel_files = []
            errors = {}
            success_files = []
            for pdf_file in pdf_files:
                pdf_file_path = os.path.join(media_dir, pdf_file.name)

                with open(pdf_file_path, 'wb') as destination:
                    for chunk in pdf_file.chunks():
                        destination.write(chunk)

                try:
                    excel_file_path = os.path.join(media_dir, f"{os.path.splitext(pdf_file.name)[0]}.xlsx")
                    excel_file_path = convert_pdf_to_excel_itkaneaswakk(pdf_file_path, excel_file_path)
                    excel_files.append(excel_file_path)
                    success_files.append(pdf_file.name)
                except ValueError as e:
                    error_message = str(e)
                    errors[pdf_file.name] = error_message

            # Si tous les fichiers PDF ont échoué
            if not success_files and errors:
                # Créer un fichier Excel pour stocker les erreurs
                error_excel = openpyxl.Workbook()
                error_sheet = error_excel.active
                # Add headers to the new sheet
                error_sheet.append(['pdf_name', 'Remarque Client/Article'])
                # Écrire les erreurs dans le fichier Excel
                # Écrire les erreurs dans le fichier Excel
                for i, (file_name, error_message) in enumerate(errors.items(), start=1):
                    error_sheet.cell(row=i + 1, column=1, value=file_name)
                    error_sheet.cell(row=i + 1, column=2, value=error_message)

                # Sauvegarder le fichier Excel
                error_report_path = os.path.join(media_dir, 'error_report.xlsx')
                error_excel.save(error_report_path)

                # Vous pouvez maintenant retourner ce fichier Excel comme réponse
                with open(error_report_path, 'rb') as excel_file:
                    response = HttpResponse(excel_file.read(),
                                            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                    response['Content-Disposition'] = 'attachment; filename=error_report.xlsx'
                    return response

            # Combine Excel files into a single Excel file
            combined_excel_file_path = os.path.join(media_dir, "combined_excel.xlsx")
            combine_excel_files(excel_files, combined_excel_file_path)

            if errors and success_files:
                # Open the combined Excel file
                combined_excel = openpyxl.load_workbook(combined_excel_file_path)

                # Remove the empty sheet if it exists (often named 'Sheet')
                if 'Sheet' in combined_excel.sheetnames:
                    sheet_to_remove = combined_excel['Sheet']
                    combined_excel.remove(sheet_to_remove)

                # Create a new sheet in 'combined_excel.xlsx' to copy data from 'error_report.xlsx'
                new_sheet = combined_excel.create_sheet(title='error_report_data')

                # Add headers to the new sheet
                new_sheet.append(['pdf_name', 'Remarque Client/Article'])

                # Open the 'error_report.xlsx' file or create it if it doesn't exist
                error_report_path = os.path.join(media_dir, 'error_report.xlsx')
                if os.path.exists(error_report_path):
                    error_report = openpyxl.load_workbook(error_report_path)
                else:
                    error_report = openpyxl.Workbook()

                # Get or create the 'error_report_data' sheet
                if 'error_report_data' not in error_report.sheetnames:
                    error_report_sheet = error_report.create_sheet(title='error_report_data')
                else:
                    error_report_sheet = error_report['error_report_data']

                # Add headers to the new sheet if it's newly created
                if error_report_sheet.max_row == 1:
                    error_report_sheet.append(['pdf_name', 'Remarque Client/Article'])

                # Add errors from the PDF files that failed to the 'error_report_data' sheet
                for pdf_file, error_message in errors.items():
                    error_report_sheet.append([pdf_file, error_message])

                # Save the 'error_report.xlsx' file
                error_report.save(error_report_path)

                # Add errors from the PDF files that failed to the 'error_report_data' sheet in the combined Excel
                for pdf_file, error_message in errors.items():
                    new_sheet.append([pdf_file, error_message])

                # Save the combined Excel file (without the empty sheet)
                combined_excel.save(combined_excel_file_path)

                # Generate the content of the text file from the combined Excel file
                text_content = generate_text_content(combined_excel_file_path)

                # Create a ZIP file
                zip_buffer = io.BytesIO()
                with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zipf:
                    # Add the combined Excel file to the ZIP
                    zipf.write(combined_excel_file_path, arcname='combined_excel_with_errors.xlsx')

                    # Add the 'error_report.xlsx' file to the ZIP
                    zipf.write(error_report_path, arcname='error_report.xlsx')

                    # Add the text file to the ZIP
                    zipf.writestr('combined_blocnote.txt', text_content)

                # Prepare the HTTP response for the ZIP file
                response = HttpResponse(zip_buffer.getvalue(), content_type='application/zip')
                response['Content-Disposition'] = 'attachment; filename=combined_files.zip'

                return response


            elif success_files:

                if remove_empty_sheet(combined_excel_file_path):
                    # Générer le contenu du fichier texte à partir du fichier Excel combiné
                    text_content = generate_text_content(combined_excel_file_path)

                    # Télécharger le fichier Excel combiné
                    excel_response = download_excel(request, combined_excel_file_path)

                    # Créer un fichier ZIP
                    zip_buffer = io.BytesIO()
                    with zipfile.ZipFile(zip_buffer, 'w') as zipf:
                        # Ajouter le fichier Excel au ZIP
                        zipf.writestr('combined_excel.xlsx', excel_response.content)
                        # Ajouter le fichier texte au ZIP
                        zipf.writestr('combined_blocnote.txt', text_content)

                    # Définir les en-têtes appropriés pour le téléchargement du fichier ZIP
                    response = HttpResponse(zip_buffer.getvalue(), content_type='application/zip')
                    response['Content-Disposition'] = 'attachment; filename=combined_files.zip'

                    return response

    return render(request, 'Myapp/index4.html')
def convert_pdf_to_excel_itkaneaswakk(input_pdf_file, output_excel_file):
    # Read the PDF and get the "No de commande" and "Nom de client"
    no_commande, nom_client = read_pdf_and_get_no_commande_client(input_pdf_file)

    if not no_commande or not nom_client:
        print("No commande or Nom de client not found in the PDF.")
        return
    a = 0
    # Create Excel Workbook
    wb = Workbook()
    # Liste pour stocker les DataFrames des feuilles valides
    valid_dfs = []
    # Initialize an empty list to store errors
    errors = []
    error_message = []
    nombre_total_de_lignes = 0
    # Read PDF File and Extract Tables
    # Read PDF File and Extract Tables
    with pdfplumber.open(input_pdf_file) as pdf:

        for i, page in enumerate(pdf.pages, start=1):
            table = page.extract_table()
            if table:
                header = table[0]  # Get the header row
                if "Article" in header:
                    combined_data = table[1:]  # Exclude header row
                    df = pd.DataFrame(combined_data, columns=header)
                    # Convert "Article" column to string type
                    df["Article"] = df["Article"].astype(str)

                    # Create a new column "UV" to store the corresponding UV value
                    df["UV"] = None

                    # Match "Article" column with "Code_Barre" in the bondedecommande table
                    articles_without_match = []
                    error_article = []
                    for index, row in df.iterrows():
                        article = row["Article"]
                        code_barre_id = bondedecommandeeee.objects.filter(Code_Barre=article).values_list('id',
                                                                                                         flat=True).first()

                        if code_barre_id is not None:
                            code_article_magasin_id = bondedecommandeeee.objects.filter(id=code_barre_id).values_list(
                                'Code_Article_magasin', flat=True).first()
                            uv = bondedecommandee.objects.filter(Code_Barre=article).values_list('UV',
                                                                                                 flat=True).first()

                            df.at[index, "Article"] = code_article_magasin_id

                            print(
                                f"Article: {article} | Code_Barre ID: {code_barre_id} | Code_Article_magasin ID: {code_article_magasin_id}")
                        else:
                            print(article)

                            error_message.append("Code_Barre not found in database")
                            error_message.append(article)
                            raise ValueError(error_message)

                    print("Articles without a match in the database:")
                    print(articles_without_match)

                    valid_dfs.append(df)

                else:
                    print(f"Table {i} does not contain the required columns.")

        # Après avoir traité tous les PDF, imprimez le nombre total de lignes
        # Ajouter le DataFrame à la liste des DataFrames valides

        #        print(df, "sanaae")
        # Save the DataFrame to the Excel sheet
        sheet_name = f"Table_{i}"
        ws = wb.create_sheet(title=sheet_name)
        for r_idx, row in enumerate(dataframe_to_rows(df, index=False), start=1):
            for c_idx, value in enumerate(row, start=1):
                ws.cell(row=r_idx, column=c_idx, value=value)

    # ...

    # Save the Excel Workbook
    wb.save(output_excel_file)
    dfs_to_concat = []  # List to store valid DataFrames for concatenation

    # Read the Excel file again to get the valid sheet names
    valid_sheet_names = [f"Table_{j}" for j in range(1, len(valid_dfs) + 1)]

    for sheet_name in valid_sheet_names:
        try:
            df = pd.read_excel(output_excel_file, sheet_name=sheet_name)

            # Vérifier si la colonne "Article" existe
            if "Article" in df.columns:
                dfs_to_concat.append(df)
            else:
                print(f"La feuille {sheet_name} ne contient pas la colonne 'Article' requise.")
        except Exception as e:
            print(f"Erreur lors de la lecture de la feuille {sheet_name} : {e}")

        # ...

        # Rest of your code for concatenating DataFrames and saving the final DataFrame

        if valid_dfs:
            combined_df = pd.concat(valid_dfs, ignore_index=True)
            combined_df = combined_df[~combined_df["Article"].astype(str).str.contains("Article")]
            print("sanaa", combined_df)
            # Assuming you have the combined_df DataFrame somewhere in your code
            # generate_report(combined_df)

            # Enregistrer le DataFrame combiné dans le fichier Excel
            combined_df.to_excel(output_excel_file, sheet_name="Combined_Table", index=False)
            print(combined_df.shape[0], "nombre de ligne de table article de ce pdf selectionner")
        else:
            print("Aucune feuille valide trouvée pour la concaténation.")

    print("All tables combined and saved to Excel.")
    print(f"No commande: {no_commande}")
    print(f"nom de client: {nom_client}")

    # Variable to store the id of the corresponding client Quanten\nUVC
    matching_client_id = None

    clients = Client.objects.all()

    # Loop through the records and print each column's value
    for client in clients:
        # Remove spaces from the "Nom_Client" field of the current client object
        nom_client_stripped = client.Nom_Client.replace(" ", "")

        # Compare the stripped Nom_Client from the PDF with the stripped Nom_Client from the database
        if nom_client_stripped.lower() == nom_client.lower():
            print("Comparison Result: True (Match)")
            # Save the id of the corresponding client
            matching_client_id = client.id

            # Display the information of the corresponding client
            print(f"ID du client correspondant: {matching_client_id}")
            print(f"Groupe_Vendeur: {client.Groupe_Vendeur}")
            print(f"Designation_Vendeur: {client.Designation_Vendeur}")
            print(f"Conditions_paiement: {client.Conditions_paiement}")
            print(f"Agence_commerciale: {client.Agence_commerciale}")
            print(f"Nom: {client.Nom_Client}")
            print(f"DIVISON: {client.DIVISON}")
            print("\n")

            # Exit the loop as we have found a match
            break

    # Check if a matching client has been found
    if matching_client_id:
        print(f"Le client correspondant a été trouvé.")

        # Add the "Groupe_Vendeur", "Designation_Vendeur", and "Conditions_paiement" columns to the "Combined_Table" DataFrame
        import datetime

        combined_df.insert(0, "Type doc vte", "ZNCS")
        combined_df.insert(1, "Organis.commerciale", "V500")
        combined_df.insert(2, "Canal_distribution", "GS")
        combined_df.insert(3, "Secteur_activite", "ZN")
        combined_df.insert(4, "Agence", client.Agence_commerciale)
        combined_df.insert(5, "Groupe de vendeurs", client.Groupe_Vendeur)

        import datetime
        today_date = datetime.date.today().strftime("%Y.%m.%d")
        # combined_df.insert(5, "Nom", client.Nom_Client)
        current_date = datetime.date.today().strftime("%Y-%m-%d")
        combined_df.insert(6, "Donneur", client.Code_Client)
        combined_df.insert(7, " Rece", client.Code_Client)
        combined_df.insert(8, "N", no_commande)
        combined_df.insert(9, "Date", today_date)
        combined_df.insert(11, "Num", "")
        combined_df.insert(13, "Itinéraire", "")
        combined_df.insert(14, "Division", client.DIVISON)
        combined_df.insert(15, "Poids", "")
        combined_df.insert(16, "TXLINE_02", "")
        combined_df.insert(17, "N équipement", "")
        # combined_df.insert(9, " Article", article)
        # combined_df.insert(10, "unité de vente", uv)
        # combined_df.insert(11, "Type poste", "TAN")
        # combined_df.insert(12, "DIVISION", client.DIVISON)

        # Save the updated DataFrame to the Excel file with the new columns
        combined_df.to_excel(output_excel_file, sheet_name="Combined_Table", index=False)

        # ...
        # combined_df["Type poste"] ="TAN"
        # combined_df["DIVISON"] = client.DIVISON

        # combined_df["Désignation"] = client.Designation_Vendeur
        # combined_df["Conditions paiement"] = client.Conditions_paiement

        # Remove the columns VL and Noligne from the DataFrame Qteenstock
        combined_df.drop(columns=["VL", "Noligne", "UV", "Libellearticle","Qteenstock","Type\nU.C."], inplace=True)

        # Remove columns that end with the word "speciale" from the DataFrame
        columns_to_drop = [col for col in combined_df.columns if col.lower().endswith("speciale")]
        combined_df.drop(columns=columns_to_drop, inplace=True)

        # Remove columns that end with the word "UC" from the DataFrame
        columns_to_drop_UC = [col for col in combined_df.columns if col.lower().endswith("UC")]
        combined_df.drop(columns=columns_to_drop_UC, inplace=True)
        # Remove columns that end with the word "UC" from the DataFrame
        # columns_to_drop_libelle = [col for col in combined_df.columns if col.lower().endswith("Libellearticle")]
        # combined_df.drop(columns=columns_to_drop_libelle, inplace=True)



        # Déplacer la colonne "quantité d'ordre" après la colonne "Article"
        if "Num" in combined_df.columns and "Quanten\nUC" in combined_df.columns:
            article_index = combined_df.columns.get_loc("Num")
            quantity_order_index = combined_df.columns.get_loc("Quanten\nUC")
            if quantity_order_index != article_index + 1:
                combined_df = combined_df[combined_df.columns[:article_index + 1].tolist() +
                                          combined_df.columns[quantity_order_index:quantity_order_index + 1].tolist() +
                                          combined_df.columns[article_index + 1:quantity_order_index].tolist() +
                                          combined_df.columns[quantity_order_index + 1:].tolist()]
        first_column_to_drop_article = next((col for col in combined_df.columns if col.lower().startswith("Num")), None)
        if first_column_to_drop_article:
            combined_df.drop(columns=first_column_to_drop_article, inplace=True)

            # column_to_replace_libelle = next((col for col in combined_df.columns if col.lower().startswith("Libelle article")), None)
            # if column_to_replace_libelle:
            #     combined_df.rename(columns={column_to_replace_libelle: "Désignation article"}, inplace=True)
        # Save the updated DataFrame to the Excel file with the new columns
        # Renommer la colonne "Article" en "Articlee"
        if "Article" in combined_df.columns:
            combined_df.rename(columns={"Article": "Articlee"}, inplace=True)
        combined_df.to_excel(output_excel_file, sheet_name="Combined_Table", index=False)
    else:
        print(f"Erreur:'{nom_client}'")
        # error_message_client = f"Le client n'a pas été trouvé dans la base de données pour '{nom_client}'"
        error_message.append("Nom client not found")
        error_message.append(nom_client)
        raise ValueError(error_message)

    # Call the generate_error_excel function to create an error Excel file
        # Load the combined DataFrame from the Excel file
    combined_df = pd.read_excel(output_excel_file, sheet_name="Combined_Table")

        # Check if the "quantite" column exists
        # Check if the "Quanten\nUC" and "UVC/UC" columns exist
    if "Quanten\nUC" in combined_df.columns and "UVC/UC" in combined_df.columns:
            combined_df["Quanten\nUC"] = combined_df["Quanten\nUC"] * combined_df["UVC/UC"]
            print("Quanten\nUC multiplied by UVC/UC")

            # Save the updated DataFrame back to the Excel file
    # Assuming you have a DataFrame called 'combined_df'
    combined_df.to_excel(output_excel_file, sheet_name="Combined_Table", index=False)
    # Check if the "UVC/UC" column exists
    if "UVC/UC" in combined_df.columns:
        combined_df.drop(columns=["UVC/UC"], inplace=True)
        print("UVC/UC column removed")

    # Save the updated DataFrame (without the "UVC/UC" column) back to the Excel file
    combined_df.to_excel(output_excel_file, sheet_name="Combined_Table", index=False)


    print("Quantite column updated and saved to Excel.")

    return output_excel_file


def generate_text_content(excel_file_path):
    # Charger le fichier Excel en utilisant pandas
    df = pd.read_excel(excel_file_path, engine='openpyxl')
    # Convertir le DataFrame en une chaîne de texte au format CSV
    #text_content = df.to_csv(index=False, header=True, sep='\t', na_rep='')
    text_content = df.to_csv(index=False, header=False, sep='\t', na_rep='')

    return text_content


def remove_empty_sheet(excel_file_path):
    try:
        # Ouvrir le fichier Excel
        wb = openpyxl.load_workbook(excel_file_path)

        # Supprimer la feuille vide (par exemple, la première feuille)
        sheet_name = wb.sheetnames[
            0]  # Vous pouvez changer cela pour correspondre à la feuille que vous souhaitez supprimer
        wb.remove(wb[sheet_name])

        # Enregistrez les modifications
        wb.save(excel_file_path)

        # Fermez le fichier Excel
        wb.close()

        return True  # Indiquez que la feuille vide a été supprimée avec succès
    except Exception as e:
        print(f"Erreur lors de la suppression de la feuille vide : {str(e)}")
        return False  # Indiquez qu'il y a eu une erreur lors de la suppression


from django.http import HttpResponse


def download_notepad(request, text_content, file_name):
    response = HttpResponse(text_content, content_type='text/plain')
    response['Content-Disposition'] = f'attachment; filename={file_name}.txt'
    return response


def download_excel(request, file_path, file_name):
    with open(file_path, 'rb') as excel_file:
        response = HttpResponse(excel_file.read(),
                                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename={file_name}'
        return response


from django.http import HttpResponse
from openpyxl import Workbook

from openpyxl import Workbook
from django.http import HttpResponse


def generate_error_excel(errors):
    # Create an Excel workbook
    workbook = Workbook()
    sheet = workbook.active

    # Add error messages to the sheet
    for i, error_message in enumerate(errors, start=1):
        sheet.cell(row=i, column=1, value=error_message)

    # Create a response with Excel content
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=error_message.xlsx'
    workbook.save(response)

    return response


import os
import openpyxl
from django.http import HttpResponse
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, landscape
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer

from openpyxl import load_workbook


def get_number_of_rows(pdf_file_path):
    num_rows = 0

    with pdfplumber.open(pdf_file_path) as pdf:
        for page in pdf.pages:
            table = page.extract_table()
            if table:
                header = table[0]
                if "Article" in header:
                    num_rows += len(table) - 1  # Exclude the header row

    return num_rows


import openpyxl


def get_number_of_rowss(excel_file_path):
    num_rows = 0

    try:
        wb = openpyxl.load_workbook(excel_file_path, data_only=True)
        ws = wb.active

        for row in ws.iter_rows(min_row=2):  # Skip the header row
            if any(cell.value for cell in row):
                num_rows += 1

    except Exception as e:
        print(f"Error reading Excel file: {e}")

    return num_rows


def combine_excel_files(excel_files, combined_excel_file_path):
    combined_wb = Workbook()

    for file_index, file_path in enumerate(excel_files):
        wb = load_workbook(file_path, data_only=True)  # Charger uniquement les valeurs (évite les erreurs de formules)

        for sheet_name in wb.sheetnames:
            if sheet_name not in combined_wb.sheetnames:
                combined_wb.create_sheet(title=sheet_name)

            source_sheet = wb[sheet_name]
            combined_sheet = combined_wb[sheet_name]

            first_row = True  # Pour suivre si c'est la première ligne de la feuille
            for row in source_sheet.iter_rows(
                    values_only=True):  # Utiliser values_only pour obtenir les valeurs des cellules
                if first_row and file_index != 0:
                    first_row = False
                    continue  # Ignorer la première ligne pour les feuilles suivantes
                combined_sheet.append(row)

    combined_wb.save(combined_excel_file_path)


def download_excel(request, file_path):
    file_name = os.path.basename(file_path)

    with open(file_path, 'rb') as excel_file:
        response = HttpResponse(excel_file.read(),
                                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="{file_name}"'
        return response


from django.shortcuts import render

from django.shortcuts import render, redirect
from .models import User

from django.urls import reverse
from django.urls import reverse

from django.contrib.auth.decorators import login_required
from django.contrib.auth import login


def commande(request):
    user = request.user  # Accéder à l'utilisateur connecté
    print("Début de la vue commande")
    # ... le reste de votre code ...
    return render(request, 'Myapp/commande.html')


from django.contrib.auth import logout
from django.shortcuts import redirect


def logout_view(request):
    logout(request)
    return redirect('login_view')  # Redirigez l'utilisateur vers la page de connexion après la déconnexion


def login_view(request):
    if request.method == 'POST':
        email = request.POST['email']
        password = request.POST['password']
        remember_me = request.POST.get('remember_me')  # Vérifiez si la case "Se souvenir de moi" est cochée.

        try:
            user = User.objects.get(email=email, password=password)

            if remember_me:
                # Marquez la session comme persistante (se souvenir de moi).
                login(request, user)
            else:
                # Utilisez la méthode de connexion normale de Django.
                request.session.set_expiry(0)  # Session de navigation.

            # Authentification réussie, redirigez vers la page souhaitée.
            if 'next' in request.GET:
                return redirect(request.GET['next'])
            else:
                return redirect('commande')  # Redirection par défaut après la connexion.
        except User.DoesNotExist:
            # L'utilisateur avec cet email et ce mot de passe n'existe pas, gérer l'erreur ici (par exemple, afficher un message d'erreur).
            pass

    # Si l'authentification échoue ou si la requête est GET sans 'next', affichez la page de connexion.
    return render(request, 'Myapp/login.html')


from django.shortcuts import render, redirect
import pandas as pd
from openpyxl import Workbook
import tempfile
import os

import pandas as pd
import tempfile
import pandas as pd
import tempfile

import pandas as pd
import tempfile

import pandas as pd
import tempfile
import pandas as pd
import tempfile

import pandas as pd
import tempfile
from django.http import HttpResponse
from django.shortcuts import render

import pandas as pd
import tempfile
from django.http import HttpResponse
from django.shortcuts import render


def comparaison_excel(request):
    if request.method == 'POST':
        # Retrieve the uploaded Excel files from the POST request
        fichiers_excel = request.FILES.getlist('pdf_file')

        # Initialize DataFrames to store data
        df_N_Article = pd.DataFrame(columns=['N', 'Articlee', 'quantité'])
        df_commande_achat_Articlee = pd.DataFrame(
            columns=['Nº commande d\'achat', 'Article', 'Quantité d\'ordre', 'Document commercial', 'Division'])

        # Read and store data from the uploaded Excel files
        for fichier in fichiers_excel:
            df = pd.read_excel(fichier)

            # Check and store data from the first Excel file
            if 'N' in df.columns and 'Articlee' in df.columns and 'quantité' in df.columns:
                df_N_Article = pd.concat([df_N_Article, df[['N', 'Articlee', 'quantité']]], ignore_index=True)

            # Check and store data from the second Excel file
            if 'Nº commande d\'achat' in df.columns and 'Article' in df.columns and 'Quantité d\'ordre' in df.columns and 'Document commercial' in df.columns:
                df_commande_achat_Articlee = pd.concat([df_commande_achat_Articlee, df[
                    ['Nº commande d\'achat', 'Article', 'Quantité d\'ordre', 'Document commercial', 'Division']]],
                                                       ignore_index=True)

        # Initialize an empty list to store matched rows
        matched_rows = []

        # Iterate through each row in df_N_Article
        # Iterate through each row in df_N_Article
        for index, row in df_N_Article.iterrows():
            N_value = row['N']
            Article_value = row['Articlee']
            quantite_value = row['quantité']

            # Check if there is a match in df_commande_achat_Articlee
            match = (df_commande_achat_Articlee['Nº commande d\'achat'] == N_value) & \
                    (df_commande_achat_Articlee['Article'] == Article_value) & \
                    (df_commande_achat_Articlee['Quantité d\'ordre'] == quantite_value)
            # Check if there is a match in df_commande_achat_Articlee
            # Check if there is a match in df_commande_achat_Articlee
            match1 = ((df_commande_achat_Articlee['Nº commande d\'achat'] == N_value) &
                      (df_commande_achat_Articlee['Article'] == Article_value) &
                      (df_commande_achat_Articlee['Quantité d\'ordre'] != quantite_value)) | \
                     ((df_commande_achat_Articlee['Nº commande d\'achat'] == N_value) &
                      (df_commande_achat_Articlee['Article'] != Article_value) &
                      (df_commande_achat_Articlee['Quantité d\'ordre'] == quantite_value))

            # If there is a match, add the row to matched_rows list with 'correspondance_N' set to True
            if match.any():
                matched_row = df_commande_achat_Articlee[match].iloc[0]
                matched_row_dict = {
                    'N': N_value,
                    'Articlee': Article_value,
                    'quantité': quantite_value,
                    'Nº commande d\'achat': matched_row['Nº commande d\'achat'],
                    'Article': matched_row['Article'],
                    'Quantité d\'ordre': matched_row['Quantité d\'ordre'],
                    'Document commercial': matched_row['Document commercial'],
                    'Division': matched_row['Division'],
                    'correspondance_N': True,  # Set 'correspondance_N' to True
                    'correspondance_Articlee': True,  # Set 'correspondance_N' to True
                    'correspondance_quantité': True,  # Set 'correspondance_N' to True
                    'Remarque': 'OK'  # Set 'correspondance_N' to True
                }
                matched_rows.append(matched_row_dict)
            elif match1.any():
                matched_row = df_commande_achat_Articlee[match1].iloc[0]
                matched_row_dict = {
                    'N': N_value,
                    'Articlee': Article_value,
                    'quantité': quantite_value,
                    'Nº commande d\'achat': matched_row['Nº commande d\'achat'],
                    'Article': matched_row['Article'],
                    'Quantité d\'ordre': matched_row['Quantité d\'ordre'],
                    'Document commercial': matched_row['Document commercial'],
                    'Division': matched_row['Division'],
                    'correspondance_N': True,  # Set 'correspondance_N' to True
                    'correspondance_Articlee': True,  # Set 'correspondance_N' to True
                    'correspondance_quantité': True,  # Set 'correspondance_N' to True
                    'Remarque': 'No OK'  # Set 'correspondance_N' to True
                }
                matched_rows.append(matched_row_dict)

            else:
                # If there is no match, print a message, and add a row with 'correspondance_N' set to False
                print(f"No match found for N={N_value}, Articlee={Article_value}, quantité={quantite_value}")
                unmatched_row_dict = {
                    'N': N_value,
                    'Articlee': Article_value,
                    'quantité': quantite_value,
                    'Nº commande d\'achat': None,
                    'Article': None,
                    'Quantité d\'ordre': None,
                    'Document commercial': None,
                    'Division': None,
                    'correspondance_N': False,  # Set 'correspondance_N' to False
                    'correspondance_Articlee': False,  # Set 'correspondance_N' to False
                    'correspondance_quantité': False,  # Set 'correspondance_N' to False
                    'Remarque': 'No ok'  # Set 'correspondance_N' to False
                }
                matched_rows.append(unmatched_row_dict)

        # Create a new DataFrame from matched_rows list
        df_combined = pd.DataFrame(matched_rows)

        # Calculate the 'ecart' column as the difference between 'quantité' and 'Quantité d'ordre'
        df_combined['ecart'] = df_combined['quantité'] - df_combined['Quantité d\'ordre']

        # Create a temporary Excel file
        with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as temp_excel:
            temp_excel_path = temp_excel.name
            df_combined.to_excel(temp_excel_path, index=False)

        # Define the response as a downloadable Excel file
        response = HttpResponse(open(temp_excel_path, 'rb').read())
        response['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        response['Content-Disposition'] = 'attachment; filename="matched_data.xlsx"'

        return response

    return render(request, 'Myapp/comparaisonExcel.html')


# convertir excel to bloc not
import os
from openpyxl import load_workbook
from django.http import HttpResponse
from django.shortcuts import render


def convert_excel_to_txt(request):
    if request.method == 'POST':
        file = request.FILES.get('excel_file')

        wb = load_workbook(file)
        sheet = wb.active

        txt_content = ""

        for row in sheet.iter_rows(values_only=True):
            row_text = "\t".join(str(cell) if cell is not None else "" for cell in row)
            txt_content += row_text + "\n"

        response = HttpResponse(content_type='text/plain')
        response['Content-Disposition'] = 'attachment; filename=excel_to_txt.txt'
        response.write(txt_content)

        return response

    return render(request, 'Myapp/convert_excel_to_txt.html')
